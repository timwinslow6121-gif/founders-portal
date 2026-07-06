import csv
import io
import json
import re
from datetime import date, datetime

import openpyxl
from dateutil.relativedelta import relativedelta
from flask import (abort, flash, redirect, render_template,
                   request, url_for, current_app, jsonify)
from flask_login import current_user, login_required
from sqlalchemy import or_

from app.extensions import db
from app.models import CommissionStatement, User, AgentCarrierContract, Policy, PolicyPayment, CommissionLineItem, AgentRecapPeriod
from app.commission import commission_bp
from app.commission.payments import build_payments, parked_payments_older_than
from app.commission.ingest import ingest_statement, compute_fingerprint, find_duplicate_statement
from app.commission.normalizers import NORMALIZERS
from app.commission.ledger import EXTRACTORS, persist_line_items, verify_statement_balance
from app.commission.recap import (build_recap, get_or_create_period, is_visible_to_agent,
                                   publish_recap, build_carrier_blocks, latest_period_with_data,
                                   all_periods_with_data, quarantined_line_items,
                                   build_aggregate_matrix, quarantine_workbench)
from app.commission.rollup import apply_rollup

# Every carrier pays Founders (the agency) directly — no carrier pays an agent
# directly. So EVERY statement is agency-level: the STATEMENT belongs to no
# single agent (agent_id=NULL) and per-row PolicyPayment attribution is resolved
# individually. This set is effectively all commission carriers; it drives the
# legacy (non-normalized) upload path, which still needs a split rate from any
# active carrier contract for statement-level expected/paid math.
AGENCY_LEVEL_CARRIERS = {"Aetna", "Devoted", "Healthspring", "Humana", "UHC", "BCBS"}


def _previous_month(today=None):
    """(label, iso) for the month BEFORE `today` — commissions pay a month behind,
    so the upload attribution defaults here. e.g. July -> ('June 2026', '2026-06')."""
    from datetime import date as _date
    today = today or _date.today()
    year, month = (today.year, today.month - 1) if today.month > 1 else (today.year - 1, 12)
    d = _date(year, month, 1)
    return d.strftime("%B %Y"), d.strftime("%Y-%m")


def _statement_date_from_sheets(carrier, sheets):
    """Best-effort statement date from the file content for carriers that embed it.
    Humana SpreadsheetML: 'CommRunDt' column on the data sheet."""
    from app.commission.payments import _parse_date
    for name, rows in (sheets or {}).items():
        if not rows:
            continue
        header = rows[0]
        # find a run/paid date column
        idx = None
        for i, h in enumerate(header):
            hl = str(h).lower()
            if hl in ("commrundt", "statement date", "payment date", "pay period"):
                idx = i; break
        if idx is not None:
            for r in rows[1:]:
                if idx < len(r) and r[idx]:
                    d = _parse_date(r[idx])
                    if d:
                        return d
    return None


def load_sheets_from_bytes(file_bytes, filename):
    """Write bytes to a temp path and load via sheet_loader (handles xlsx/xls/SpreadsheetML)."""
    import tempfile, os as _os
    suffix = ".xlsx" if (filename or "").lower().endswith("xlsx") else ".xls"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tf:
        tf.write(file_bytes)
        tmp = tf.name
    try:
        from app.commission.sheet_loader import load_sheets
        return load_sheets(tmp)
    finally:
        _os.unlink(tmp)


def _detect_carrier_from_headers(header_cells):
    """Same fingerprints as _detect_carrier, but on a plain list of header cells."""
    headers = [str(c or "").lower() for c in header_cells]
    header_str = " ".join(headers)
    if "commission action" in header_str and "writing agent" in header_str:
        return "UHC"
    if "payee amount" in header_str and "sales event" in header_str:
        return "Aetna"
    if "commrundt" in header_str or "grpname" in header_str:
        return "Humana"
    if "billed amount" in header_str or ("group type" in header_str and "customer name" in header_str):
        return "BCBS"
    if "member hicn" in header_str or "agent npn" in header_str:
        return "Devoted"
    if "payment type" in header_str and "medicare beneficiary identifier" in header_str:
        return "Healthspring"
    if "distributor number" in header_str and "advance type" in header_str:
        return "Wellable"
    return None


def _detect_carrier_from_sheets(sheets):
    """Scan every sheet's first row for a recognizable carrier header.

    Multi-sheet carrier files (Devoted, Healthspring) keep their data on a
    non-first/non-active sheet, so the legacy active-sheet-only detector misses
    them. This scans all sheets and returns the first carrier matched.
    """
    for rows in sheets.values():
        if not rows:
            continue
        carrier = _detect_carrier_from_headers(rows[0])
        if carrier:
            return carrier
    return None


def _resolve_agent_from_facts(facts):
    """Pick the first writing-agent name present in the normalized facts and
    match it to a portal User. Returns user id or None (agency-level)."""
    for f in facts:
        raw = (getattr(f, "writing_agent_raw", "") or "").strip()
        if raw:
            return _match_agent_name(raw)
    return None


def _gross_preview(facts):
    return round(sum(f.amount for f in facts if f.amount > 0), 2)


def _scan_summary(ws):
    """
    Scan every cell in the sheet for AJ's manually typed summary rows.
    AJ places these inconsistently — column and row position varies month to month.

    Looks for:
      - Gross×rate row:  "7,566.59 x.55"  or  "$202.44 x.525"  etc.
      - Paid row:        a numeric cell adjacent to or near a gross×rate cell,
                         OR a cell matching "$N + $N" pattern (UHC style)

    Returns:
      (paid, stated_rate)
      paid        — the numeric amount AJ says was paid (float or None)
      stated_rate — the split rate AJ used in his formula (float or None)
                    Callers should compare this against the contract rate.
    """
    all_cells = []
    for row in ws.iter_rows():
        for cell in row:
            all_cells.append(cell)

    paid = None
    stated_rate = None

    for cell in all_cells:
        val = str(cell.value or "").strip()

        # Pattern: "NNN x .55" or "$NNN,NNN.NN x.525" — gross × split summary
        m = re.search(r'[\$]?([\d,]+\.?\d*)\s*x\.?\s*(\.?\d+)', val)
        if m:
            try:
                rate = float(m.group(2))
                if 0 < rate < 1:
                    stated_rate = rate
                elif rate > 1:          # e.g. "x55" instead of "x.55"
                    stated_rate = rate / 100
            except ValueError:
                pass
            # The paid value is often the numeric cell immediately to the right
            # or below this cell
            right = ws.cell(row=cell.row, column=cell.column + 1)
            below = ws.cell(row=cell.row + 1, column=cell.column)
            for candidate in (right, below):
                if isinstance(candidate.value, (int, float)):
                    paid = float(candidate.value)
                    break
            continue

        # Pattern: "$4,161.62 + $130.81" — paid = numeric in next cell
        if re.search(r'^\$[\d,]+\.\d+\s*\+\s*[\$\d]', val):
            right = ws.cell(row=cell.row, column=cell.column + 1)
            below = ws.cell(row=cell.row + 1, column=cell.column)
            for candidate in (right, below):
                if isinstance(candidate.value, (int, float)):
                    if paid is None:   # don't overwrite if already found
                        paid = float(candidate.value)
                    break
            continue

        # Pattern: "$283.17 + 27(last month)" — free-text paid note, extract numeric
        m2 = re.search(r'^\$?([\d,]+\.\d{2})\s*\+\s*[\d\$]', val)
        if m2:
            try:
                if paid is None:
                    paid = float(m2.group(1).replace(',', ''))
            except ValueError:
                pass

    return paid, stated_rate


def _parse_uhc(ws):
    # April 2026 UHC column layout (0-indexed):
    #  col3:  Statement Date     col5:  Writing Agent Name
    #  col7:  Member Name        col8:  MedicareID (MBI)
    #  col11: Original Eff Date  col19: Commission Action
    #  col23: Commission         col24: Term Reason    col28: Term Date
    paid, stated_rate = _scan_summary(ws)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    line_items = []
    gross = 0.0
    bonus = 0.0
    stmt_date = None

    for row in rows:
        if not any(row):
            continue
        action     = str(row[19] or "").strip()
        commission = row[23]

        if stmt_date is None and row[3] and isinstance(row[3], datetime):
            stmt_date = row[3].date()

        if action in ("New Chargeback",):
            amt = float(commission) if commission and isinstance(commission, (int, float)) else None
            if amt:
                gross += amt
            line_items.append({
                "mbi":         str(row[8] or "").strip(),
                "member":      str(row[7] or ""),
                "eff_date":    str(row[11].date() if isinstance(row[11], datetime) else row[11] or ""),
                "action":      action,
                "amount":      amt,
                "term_reason": str(row[24] or ""),
            })
            continue

        if action in ("Renewal", "New"):
            amt = float(commission) if commission and isinstance(commission, (int, float)) else None
            if amt:
                gross += amt
            line_items.append({
                "mbi":         str(row[8] or "").strip(),
                "member":      str(row[7] or ""),
                "eff_date":    str(row[11].date() if isinstance(row[11], datetime) else row[11] or ""),
                "action":      action,
                "amount":      amt,
                "term_reason": str(row[24] or ""),
            })

    return gross, bonus, paid or 0.0, stmt_date, line_items, stated_rate


def _parse_aetna(ws):
    # April 2026 Aetna CSV column layout (0-indexed):
    #  col0:  Payment Date       col1:  Medicare Number (MBI)
    #  col4:  Member Name        col6:  Sales Event (action)
    #  col9:  Plan ID            col12: Coverage Period
    #  col14: Writing Agent NPN  col16: Writing Agent Name
    #  col20: Payee Amount
    paid, stated_rate = _scan_summary(ws)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    line_items = []
    gross = 0.0
    stmt_date = None

    for row in rows:
        if not any(row):
            continue
        # Skip footer rows (e.g. "Total Payee Amount: $1461.97" in col0)
        col0 = str(row[0] or "").strip()
        if col0.lower().startswith("total"):
            continue
        # Skip summary rows
        if re.search(r'[\d,]+\.?\d*\s*x', str(row[9] or "")):
            continue

        amount = row[20] if len(row) > 20 else None   # Payee Amount
        mbi    = str(row[1] or "").strip()             # Medicare Number

        # Parse stmt_date from Payment Date (col0) — may be string "YYYY-MM-DD"
        if stmt_date is None and row[0]:
            if isinstance(row[0], datetime):
                stmt_date = row[0].date()
            elif isinstance(row[0], date):
                stmt_date = row[0]
            else:
                try:
                    stmt_date = datetime.strptime(str(row[0]).strip(), "%Y-%m-%d").date()
                except ValueError:
                    pass

        # Convert string amounts (CSV comes in as strings)
        if isinstance(amount, str):
            try:
                amount = float(amount.replace(",", "").replace("$", "").strip())
            except ValueError:
                amount = None

        if amount and isinstance(amount, (int, float)) and float(amount) != 0:
            gross += float(amount)
            line_items.append({
                "mbi":      mbi,
                "member":   str(row[4] or ""),
                "plan":     str(row[9] or ""),
                "eff_date": str(row[12] or ""),
                "action":   str(row[6] or ""),
                "amount":   float(amount),
            })

    return gross, 0.0, paid or 0.0, stmt_date, line_items, stated_rate


def _parse_humana(ws):
    paid_scan, stated_rate = _scan_summary(ws)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    line_items = []
    gross     = 0.0
    stmt_date = None

    for row in rows:
        if not any(row):
            continue
        # Skip summary rows
        if re.search(r'[\$\d,]+\.?\d*\s*x', str(row[8] or "")):
            continue
        if re.search(r'[\$\d,]+\.?\d*\s*x', str(row[7] or "")):
            continue

        amount  = row[8]   # PaidAmount
        comment = str(row[9] or "").strip()

        if stmt_date is None and row[1] and isinstance(row[1], datetime):
            stmt_date = row[1].date()

        if amount and isinstance(amount, (int, float)):
            gross += float(amount)
            line_items.append({
                "member":  str(row[4] or ""),
                "month":   str(row[6] or ""),
                "action":  comment,
                "amount":  float(amount),
                "product": str(row[7] or ""),
            })

    # Humana (like every carrier) pays Founders; the statement is agency-level.
    # Use scanned paid if available, otherwise gross.
    paid = paid_scan if paid_scan is not None else gross
    return gross, 0.0, paid, stmt_date, line_items, stated_rate


def _parse_bcbs(ws):
    paid, stated_rate = _scan_summary(ws)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    line_items = []
    gross = 0.0
    stmt_date = None  # will be set from file content or filename fallback

    for row in rows:
        if not any(row):
            continue
        # Skip summary rows
        if re.search(r'[\$\d,]+\.?\d*\s*x', str(row[9] or "")):
            continue
        col13 = str(row[13] or "").strip() if len(row) > 13 else ""
        if col13.startswith("="):
            continue
        col12 = str(row[12] or "").strip() if len(row) > 12 else ""
        if col12.lower() == "total:":
            continue

        commission = row[13] if len(row) > 13 else None
        if commission and isinstance(commission, (int, float)) and float(commission) != 0:
            gross += float(commission)
            line_items.append({
                "member":     str(row[3] or ""),
                "plan":       str(row[6] or ""),
                "group_type": str(row[2] or ""),
                "eff_date":   str(row[5] or ""),
                "action":     str(row[2] or ""),
                "amount":     float(commission),
            })

    return gross, 0.0, paid or 0.0, stmt_date, line_items, stated_rate


def _parse_devoted(ws):
    paid, stated_rate = _scan_summary(ws)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    line_items = []
    gross = 0.0
    stmt_date = None

    for row in rows:
        if not any(row):
            continue
        # Skip summary rows
        if re.search(r'[\$\d,]+\s*x\.?\s*\.?\d+', str(row[8] or "")):
            continue

        amount = row[11]  # Base Amount
        if amount and isinstance(amount, (int, float)):
            gross += float(amount)
            line_items.append({
                "member":    f"{row[5] or ''} {row[6] or ''}".strip(),
                "member_id": str(row[3] or ""),
                "eff_date":  str(row[7] or ""),
                "period":    str(row[10] or ""),
                "action":    str(row[9] or "New/Renewal"),
                "amount":    float(amount),
            })

        if stmt_date is None and row[0]:
            try:
                stmt_date = datetime.strptime(str(row[0]), "%m/%d/%Y").date()
            except Exception:
                pass

    return gross, 0.0, paid or 0.0, stmt_date, line_items, stated_rate


def _parse_healthspring(ws):
    paid, stated_rate = _scan_summary(ws)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    line_items = []
    gross = 0.0
    stmt_date = None

    for row in rows:
        if not any(row):
            continue
        # Skip summary rows
        if re.search(r'[\d,]+\s*x\.?\s*\.?\d+', str(row[6] if len(row) > 6 else "")):
            continue

        amount = row[7] if len(row) > 7 else None
        if amount and isinstance(amount, (int, float)):
            gross += float(amount)
            pay_period = row[6]
            if stmt_date is None and isinstance(pay_period, datetime):
                stmt_date = pay_period.date()
            line_items.append({
                "member":      str(row[8] or ""),
                "mbi":         str(row[9] or ""),
                "action":      str(row[0] or ""),
                "description": str(row[1] or ""),
                "amount":      float(amount),
            })

    return gross, 0.0, paid or 0.0, stmt_date, line_items, stated_rate


def _parse_wellable(ws):
    """Wellable advance commissions — flagged as clawback-eligible advances."""
    paid, stated_rate = _scan_summary(ws)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    line_items = []
    gross = 0.0
    stmt_date = None

    for row in rows:
        if not any(row):
            continue
        # Skip summary rows
        if re.search(r'[\$\d,]+\.?\d*\s*x\s*\.?\d+', str(row[16] if len(row) > 16 else "")):
            continue

        advance_amount = row[16] if len(row) > 16 else None
        if advance_amount and isinstance(advance_amount, (int, float)):
            gross += float(advance_amount)
            app_date = row[17] if len(row) > 17 else None
            if stmt_date is None and isinstance(app_date, datetime):
                stmt_date = app_date.date()
            line_items.append({
                "member":         str(row[5] or ""),
                "policy":         str(row[4] or ""),
                "plan":           str(row[7] or ""),
                "premium":        float(row[12]) if row[12] else 0.0,
                "advance_pct":    float(row[13]) if row[13] else 0.0,
                "advance_months": str(row[14] or ""),
                "action":         str(row[15] or ""),
                "amount":         float(advance_amount),
                "is_advance":     True,
            })

    return gross, 0.0, paid or 0.0, stmt_date, line_items, stated_rate


def _csv_bytes_to_workbook(file_bytes):
    """Convert a CSV file (bytes) into an openpyxl Workbook so the rest of
    the commission pipeline can treat it identically to an XLSX file.
    Values are stored as strings; _parse_aetna handles string-to-float conversion."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in reader:
        ws.append(row)
    return wb


_MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _parse_date_from_filename(filename):
    """
    Extract a statement month/year from a filename as a last-resort fallback.
    Handles patterns like:
      UHC_March_2026.xlsx
      BCBS April 2026 Commission.xlsx
      Humana-2026-03.xlsx
      statement_2026_04.csv
    Returns date(year, month, 1) or None if no match.
    """
    if not filename:
        return None
    name = filename.lower()

    # Pattern 1: named month + 4-digit year  e.g. "march_2026", "april 2026"
    m = re.search(
        r'(' + '|'.join(_MONTH_NAMES.keys()) + r')[_\-\s]+(\d{4})',
        name
    )
    if m:
        month = _MONTH_NAMES[m.group(1)]
        year = int(m.group(2))
        if 2020 <= year <= 2099:
            return date(year, month, 1)

    # Pattern 2: 4-digit year + named month  e.g. "2026_march"
    m = re.search(
        r'(\d{4})[_\-\s]+(' + '|'.join(_MONTH_NAMES.keys()) + r')',
        name
    )
    if m:
        year = int(m.group(1))
        month = _MONTH_NAMES[m.group(2)]
        if 2020 <= year <= 2099:
            return date(year, month, 1)

    # Pattern 3: YYYY-MM or YYYY_MM  e.g. "2026-03", "2026_04"
    m = re.search(r'(\d{4})[_\-](\d{2})', name)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        if 2020 <= year <= 2099 and 1 <= month <= 12:
            return date(year, month, 1)

    return None


def _carrier_supported_or_reason(carrier):
    """Return (True, '') if we can ingest this carrier, else (False, reason)."""
    from app.commission.normalizers import NORMALIZERS
    if not carrier:
        return False, ("Could not detect the carrier from this file's headers. "
                       "Nothing was imported.")
    if carrier not in NORMALIZERS:
        supported = ", ".join(sorted(NORMALIZERS))
        return False, (f"Cannot parse this file — carrier '{carrier}' is not yet "
                       f"supported (supported: {supported}). Nothing was imported.")
    return True, ""


def _detect_carrier(ws):
    headers = [str(c.value or "").lower() for c in ws[1]]
    header_str = " ".join(headers)
    if "commission action" in header_str and "writing agent" in header_str:
        return "UHC"
    if "payee amount" in header_str and "sales event" in header_str:
        return "Aetna"
    if "commrundt" in header_str or "grpname" in header_str:
        return "Humana"
    if "billed amount" in header_str or ("group type" in header_str and "customer name" in header_str):
        return "BCBS"
    if "member hicn" in header_str or "agent npn" in header_str:
        return "Devoted"
    if "payment type" in header_str and "medicare beneficiary identifier" in header_str:
        return "Healthspring"
    if "distributor number" in header_str and "advance type" in header_str:
        return "Wellable"
    return None


def _normalize_name(s):
    """Normalize agent name for fuzzy matching.
    Handles formats:
      - 'WINSLOW, TIMOTHY JAMES' → 'timothy winslow'
      - 'WINSLOW TIMOTHY J'      → 'timothy winslow'
      - 'Timothy Winslow'        → 'timothy winslow'
    """
    s = str(s or "").strip().lower()

    if "," in s:
        # "WINSLOW, TIMOTHY JAMES" → ["winslow", "timothy james"]
        parts = [p.strip() for p in s.split(",", 1)]
        last  = parts[0].strip()
        first = parts[1].strip().split()[0] if parts[1].strip() else ""
        return f"{first} {last}".strip()

    words = s.split()
    if len(words) == 1:
        return s
    if len(words) == 2:
        # "timothy winslow" — already normalized
        return s
    # 3+ words: could be "WINSLOW TIMOTHY J" (last first initial)
    # or "Timothy James Winslow" (first middle last)
    # Check if first word looks like a last name by seeing if it matches
    # any known last name pattern — simplest: try both orderings
    # Return "first last" by taking word[1] word[0] (last-first-initial format)
    # This handles Humana's "WINSLOW TIMOTHY J" → "timothy winslow"
    return f"{words[1]} {words[0]}".strip()


def _detect_agent_id(ws, carrier):
    """Extract agent name from file and match to a User in the database."""
    agent_col_map = {
        "UHC":          5,   # Writing Agent Name (col F, index 5)
        "Aetna":       16,   # Writing Agent Name (col Q, index 16)
        "Humana":       2,   # WaName (col C, index 2)
        "BCBS":         1,   # Agent Name (col B, index 1)
        "Devoted":      2,   # Agent Name (col C, index 2)
        "Healthspring": 3,   # Writing Broker Name (col D, index 3)
        "Wellable":     3,   # Writing Agent Name (col D, index 3)
    }
    col_idx = agent_col_map.get(carrier)
    if col_idx is None:
        return None

    # Find first non-empty agent name in data rows
    agent_name_raw = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[col_idx] if len(row) > col_idx else None
        if val and str(val).strip():
            agent_name_raw = str(val).strip()
            break

    if not agent_name_raw:
        return None

    return _match_agent_name(agent_name_raw)


_NICKNAMES = {
    "michael": "mike", "mike": "michael",
    "christopher": "chris", "chris": "christopher",
    "timothy": "tim", "tim": "timothy",
    "william": "bill", "bill": "william",
    "robert": "bob", "bob": "robert",
    "richard": "rick", "rick": "richard",
    "james": "jim", "jim": "james",
    "thomas": "tom", "tom": "thomas",
}


def _ledger_split_lookup(writing_agent_raw, carrier, agency_id):
    """Split rate for a writing agent on a carrier, snapshotted at import.
    Retired-agent (Cyndi/Don) Aetna/UHC business rolls up to Brian first, so the
    rate comes from Brian's 0.50 contract. Falls back to any active contract for
    the carrier, then 0.55.

    Takes agency_id EXPLICITLY — never reads the global current_user. current_user
    is bound only inside a request; the ingest (and scripts/re-imports) call this
    with just an app context, where current_user is unavailable and reading
    current_user.agency_id crashed the whole upload with 'NoneType has no
    attribute agency_id'."""
    writing_agent_raw = apply_rollup(writing_agent_raw, carrier)
    agent_id = _match_agent_name(writing_agent_raw) if writing_agent_raw else None
    contract = None
    if agent_id:
        contract = AgentCarrierContract.query.filter_by(
            agent_id=agent_id, carrier=carrier, is_active=True,
            agency_id=agency_id).first()
    if contract is None:
        contract = AgentCarrierContract.query.filter_by(
            carrier=carrier, is_active=True,
            agency_id=agency_id).first()
    return contract.split_rate if contract else 0.55


# Maiden-/legal-name aliases that the last-name fuzzy matcher can't bridge.
# Keyed on _normalize_name() output ("first last"). Betty Marlowe writes some
# carrier business under her legal name "RIDDLE, BETTY B" → "betty riddle".
_NAME_ALIASES = {
    "betty riddle": "betty marlowe",
}


def _match_agent_name(agent_name_raw):
    """Match a raw 'Last, First' / 'First Last' agent name to a portal User id."""

    def _first_matches(a, b):
        return a == b or _NICKNAMES.get(a) == b or _NICKNAMES.get(b) == a

    users = User.query.all()

    def _match_normalized(normalized):
        """Match an already-normalized 'first last' string to exactly one user id,
        using exact-normalized then nickname/fuzzy (same last name + first-3) logic.
        Returns the user id, or None on no match / ambiguity."""
        if not normalized:
            return None
        # Maiden-/legal-name alias (e.g. "betty riddle" → "betty marlowe").
        normalized = _NAME_ALIASES.get(normalized, normalized)
        # Exact match after normalisation
        for user in users:
            if _normalize_name(user.name) == normalized:
                return user.id
        # Fuzzy: same last name + first name matches via nickname or prefix
        np = normalized.split()
        for user in users:
            up = _normalize_name(user.name).split()
            if len(np) >= 2 and len(up) >= 2 and np[-1] == up[-1]:
                if _first_matches(np[0], up[0]) or np[0][:3] == up[0][:3]:
                    return user.id
        return None

    # 1) Straightforward normalization (handles "Last, First", "Last First MI",
    #    and already "First Last").
    matched = _match_normalized(_normalize_name(agent_name_raw))
    if matched is not None:
        return matched

    # 2) Ambiguous-order fallback: from word count alone you cannot tell
    #    "LONG REBEKAH" (last first) from "REBEKAH LONG" (first last). If the raw
    #    name has exactly two tokens (ignoring a trailing single-letter middle
    #    initial), try BOTH orderings against the user list and accept whichever
    #    resolves to a real user.
    raw = str(agent_name_raw or "").strip()
    if "," not in raw:
        tokens = raw.lower().split()
        # Drop a trailing single-letter middle initial (e.g. "FREEMAN BRIAN L").
        if len(tokens) == 3 and len(tokens[2]) == 1:
            tokens = tokens[:2]
        if len(tokens) == 2:
            for first, last in ((tokens[0], tokens[1]), (tokens[1], tokens[0])):
                m = _match_normalized(f"{first} {last}")
                if m is not None:
                    return m

    return None


PARSERS = {
    "UHC":         _parse_uhc,
    "Aetna":       _parse_aetna,
    "Humana":      _parse_humana,
    "BCBS":        _parse_bcbs,
    "Devoted":     _parse_devoted,
    "Healthspring": _parse_healthspring,
    "Wellable":    _parse_wellable,
}


def _enrich_line_items_with_commission_type(statements, agency_id):
    """
    For each statement's line items, look up commission_type from Policy by MBI.
    Mutates items in place, adding 'commission_type' key where a match exists.
    """
    all_mbis = set()
    for s in statements:
        for item in s.line_items_parsed:
            mbi = item.get('mbi', '').strip()
            if mbi:
                all_mbis.add(mbi)
    if not all_mbis:
        return
    rows = (Policy.query
            .filter(Policy.mbi.in_(all_mbis), Policy.agency_id == agency_id)
            .with_entities(Policy.mbi, Policy.commission_type)
            .all())
    mbi_to_type = {r.mbi: r.commission_type for r in rows if r.commission_type}
    for s in statements:
        for item in s.line_items_parsed:
            mbi = item.get('mbi', '').strip()
            if mbi and mbi in mbi_to_type:
                item['commission_type'] = mbi_to_type[mbi]


@commission_bp.route("/commissions")
@login_required
def commission_index():
    statements = (CommissionStatement.query
                  .filter_by(agent_id=current_user.id, agency_id=current_user.agency_id)
                  .order_by(CommissionStatement.statement_date.desc())
                  .all())
    for s in statements:
        s.line_items_parsed = json.loads(s.line_items) if s.line_items else []
    _enrich_line_items_with_commission_type(statements, current_user.agency_id)
    return render_template("commission.html",
        statements=statements, is_admin=False, viewing_agent=None)


@commission_bp.route("/admin/commissions")
@login_required
def commission_admin():
    if not current_user.is_admin:
        abort(403)
    from app.commission.recap import (commission_audit_overview, all_periods_with_data,
                                       latest_period_with_data)
    agency_id = current_user.agency_id
    # Period: ?period=, else the latest period with data, else the current month.
    periods = all_periods_with_data(agency_id) or []
    period = (request.args.get("period") or latest_period_with_data(agency_id)
              or date.today().strftime("%B %Y"))
    # Always offer the current month in the picker (so AJ can target it before any upload).
    current_month = date.today().strftime("%B %Y")
    if current_month not in periods:
        periods = [current_month] + periods
    overview = commission_audit_overview(agency_id, period)
    # Recent uploads (across all periods) — keep the delete-an-upload affordance.
    recent = (CommissionStatement.query
              .filter_by(agency_id=agency_id)
              .order_by(CommissionStatement.upload_date.desc())
              .limit(20).all())
    # Default upload month is the previous month (commissions pay a month behind).
    default_upload_month, default_upload_month_iso = _previous_month()
    return render_template("commission.html",
        overview=overview, periods=periods, selected_period=period,
        current_month=current_month, current_month_iso=date.today().strftime("%Y-%m"),
        default_upload_month=default_upload_month, default_upload_month_iso=default_upload_month_iso,
        recent=recent, is_admin=True, viewing_agent=None)


def _ingest_normalized_upload(carrier, sheets, file_bytes, filename, statement_month,
                              agency_id, actor, replace=False):
    """Run the normalize→resolve→pay pipeline for a clean-split carrier.

    Handles statement-period resolution, agent/contract/split resolution, the
    fingerprint duplicate guard, statement upsert, ingest, and summary.

    TRANSACTION OWNERSHIP: this function calls db.session.flush() — NOT commit().
    The caller (_process_one_file / commission_upload) owns the commit so that a
    db.session.begin_nested() savepoint in a multi-file loop can roll back only
    one failed file without affecting the others.

    Parameters
    ----------
    carrier:         detected carrier string (must be in NORMALIZERS)
    sheets:          dict of sheet-name → rows from load_sheets_from_bytes()
    file_bytes:      raw uploaded bytes (used for file_scoped replace)
    filename:        original filename (for period auto-detection + Healthspring token)
    statement_month: "YYYY-MM" string from the caller (form field or None/empty)
    agency_id:       the acting agency's id (NOT read from current_user inside here)
    actor:           the uploading User object (or None in tests); used for
                     uploaded_by_id and agency-scoped contract lookups
    replace:         if True, overwrite an existing duplicate statement (dup-guard bypass)

    Returns a result dict:
      success → {"filename","ok":True,"carrier","scope","rows","gross","period",
                  "warnings": [...]}
      failure → {"filename","ok":False,"error":<str>,"fix":<str|None>}
    UHC never reaches here (not in NORMALIZERS).
    """
    # Record the uploaded filename so multi-batch carriers (Healthspring) can derive
    # a per-file token from it during normalize/extract (see ledger._healthspring_filetoken).
    from app.commission.ledger import current_upload_filename
    current_upload_filename.set(filename or "")

    from app.commission.normalizers import BcbsColumnError
    try:
        facts = NORMALIZERS[carrier](sheets)
    except BcbsColumnError as e:
        # A required column is missing because the carrier/Tidewater format changed.
        # Surface the specific reason to AJ instead of a vague error or a 500.
        return {
            "filename": filename, "ok": False,
            "error": str(e),
            "fix": "check the column names / re-export from Tidewater",
        }
    if not facts:
        return {
            "filename": filename, "ok": False,
            "error": f"No commission rows found in the {carrier} file.",
            "fix": None,
        }

    # Collect non-fatal warnings to surface to the caller (previously flashed).
    warnings = []

    # Statement period: caller override → filename → file content → today.
    stmt_date = None
    form_month = (statement_month or "").strip()  # "YYYY-MM"
    if form_month:
        try:
            stmt_date = datetime.strptime(form_month, "%Y-%m").date()
        except ValueError:
            stmt_date = None
    if not stmt_date:
        stmt_date = _parse_date_from_filename(filename)
    if not stmt_date:
        stmt_date = _statement_date_from_sheets(carrier, sheets)
    if not stmt_date:
        stmt_date = date.today()
        warnings.append(
            "Could not detect statement period from the file or filename. "
            "Defaulted to today's month. Use the 'Statement Month' field to correct this.")
    period_label = stmt_date.strftime("%B %Y")

    # Devoted ships two files per month under ONE (carrier, period) statement.
    # They only coexist if both land in the SAME period_label. Rebekah's per-agent
    # statement file has no month in its filename, so if AJ didn't set the
    # Statement Month explicitly, warn to confirm it matches the agency file —
    # otherwise the two files silently split into separate statements.
    if carrier == "Devoted" and not form_month:
        from app.commission.ledger import _devoted_format
        try:
            if _devoted_format(sheets) == "statement":
                warnings.append(
                    f"This is Devoted's per-agent statement file. It was filed under "
                    f"{period_label} (auto-detected). Confirm this matches the month "
                    f"of the agency Devoted file — set 'Statement Month' on both uploads "
                    f"so they combine into one statement.")
        except ValueError:
            pass

    # Agent + split. These carriers are often agency-level (multiple writing
    # agents in one file); attribute the statement to the first writing agent
    # found, but resolve the split from any active contract for the carrier.
    agent_id = _resolve_agent_from_facts(facts)
    contract = None
    if agent_id:
        contract = AgentCarrierContract.query.filter_by(
            agent_id=agent_id, carrier=carrier, is_active=True,
            agency_id=agency_id).first()
    if contract is None:
        contract = AgentCarrierContract.query.filter_by(
            carrier=carrier, is_active=True,
            agency_id=agency_id).first()
    agent_split = contract.split_rate if contract else 0.55

    # Do NOT attribute unresolved rows to the uploading admin — that wrongly made
    # the uploader (AJ) the agent for every stub the file couldn't resolve. When no
    # writing agent matches a portal user, leave agent_id None: the row resolves to
    # an UNASSIGNED customer (primary_agent_id NULL, no AOR interval) until someone
    # sets the real agent in the portal. (See _create_stub / _open_aor_interval.)
    # agent_id stays None here on purpose.

    # Every carrier pays Founders (the agency), never an agent directly. The
    # statement is always agency-level; per-agent earnings come from per-row
    # PolicyPayment splits (resolved via agent_resolver). agent_id stays the
    # per-row fallback only.
    statement_agent_id = None

    fingerprint = compute_fingerprint(carrier, period_label, facts)
    # replace is passed in as a parameter — do NOT read request.form here; this
    # function must be fully request-context-free so it works inside a multi-file
    # loop (begin_nested savepoints) and in tests without a live request.
    # Duplicate = same content in the SAME period. A byte-identical per-agent
    # statement in a DIFFERENT pay period (BCBS/Tidewater steady-state books) is a
    # legitimate new statement, not a re-upload — see find_duplicate_statement.
    dup = find_duplicate_statement(agency_id, carrier, fingerprint, period_label)
    if dup is not None and not replace:
        return {
            "filename": filename, "ok": False,
            "error": (
                f"This looks like the {carrier} {dup.period_label} statement already "
                f"imported on {dup.statement_date:%b %d, %Y} "
                f"({len(facts)} members, ${_gross_preview(facts):,.2f}). "
                f"No payments were created. Re-submit with 'Replace existing' to overwrite."
            ),
            "fix": "re-submit with Replace existing checked",
        }

    existing = CommissionStatement.query.filter_by(
        carrier=carrier, agent_id=statement_agent_id, period_label=period_label,
        agency_id=agency_id).first()
    stmt = existing or CommissionStatement(
        carrier=carrier, agent_id=statement_agent_id, agency_id=agency_id)
    if not existing:
        db.session.add(stmt)
    stmt.statement_date = stmt_date
    stmt.period_label = period_label
    stmt.split_rate = agent_split
    stmt.filename = filename
    stmt.uploaded_by_id = actor.id if actor else None
    stmt.content_fingerprint = fingerprint
    db.session.flush()

    # If replacing, clear stale ledger rows so re-ingest doesn't double-count
    # rows that no longer appear in the file.
    if existing:
        # Replace-on-reupload. Agency-wide carriers (one file/month) blanket-delete
        # the statement's rows. Per-agent carriers (BCBS = one file per agent;
        # Devoted = agency + Rebekah) ship MULTIPLE files under one statement — scope
        # the delete to JUST the uploaded file's rows (by source_ref token) so the
        # other agents'/files' line items survive. Otherwise uploading agent B's
        # file silently wipes agent A's data.
        from app.commission.ledger import file_scoped_prefix
        pp_q = PolicyPayment.query.filter_by(
            statement_id=stmt.id, agency_id=agency_id)
        li_q = CommissionLineItem.query.filter_by(
            statement_id=stmt.id, agency_id=agency_id)
        prefix = file_scoped_prefix(carrier, sheets)
        if prefix is not None:
            pp_q = pp_q.filter(PolicyPayment.source_ref.like(prefix))
            li_q = li_q.filter(CommissionLineItem.source_ref.like(prefix))
        pp_q.delete(synchronize_session=False)
        li_q.delete(synchronize_session=False)
        db.session.flush()

    # Resolve writing-agent names to portal users AFTER retired-agent rollup, so
    # Cyndi/Don Aetna/UHC rows attribute to Brian (matching the split-rate seam).
    def _rollup_resolver(raw, c=carrier):
        return _match_agent_name(apply_rollup(raw, c))

    try:
        ingest = ingest_statement(stmt, carrier, agent_id, agency_id, sheets,
                                  agent_resolver=_rollup_resolver)

        # R1 ledger: persist EVERY sheet row (incl. Founders overrides the
        # customer-sync normalizer collapses away) so the balance is provable.
        extractor, _money = EXTRACTORS.get(carrier, (None, None))
        if extractor is not None:
            drafts = extractor(sheets, split_lookup=lambda raw, c=carrier: _ledger_split_lookup(raw, c, agency_id))
            persist_line_items(carrier, drafts, stmt, agency_id,
                               agent_resolver=_rollup_resolver)
            db.session.flush()
            report = verify_statement_balance(carrier, drafts, sheets)
            # A3 — persist the balance result so it's a VISIBLE status, not just a log.
            stmt.balanced = bool(report.completeness_ok and report.internal_ok)
            stmt.ledger_total = report.lineitem_total
            stmt.money_rows_total = report.money_rows_total
            if not report.completeness_ok:
                current_app.logger.warning(
                    "Commission ledger completeness check FAILED for "
                    f"{carrier} {period_label}: {report}")
            if not report.internal_ok:
                # internal balance is true by construction; a failure here means
                # a float-precision or derivation bug — surface it loudly.
                current_app.logger.warning(
                    "Commission ledger INTERNAL balance failed (unexpected) for "
                    f"{carrier} {period_label}: {report}")

        stmt.gross_amount = round(sum(f.amount for f in facts if f.amount > 0), 2)
        stmt.bonus_amount = 0.0
        stmt.expected_amount = round(stmt.gross_amount * agent_split, 2)
        stmt.paid_amount = stmt.expected_amount
        stmt.difference = 0.0
        stmt.status = "verified"
        # NOTE: db.session.flush() here — NOT commit(). Transaction ownership moved
        # to the caller (commission_upload or the multi-file loop in task 6) so that
        # a begin_nested() savepoint can roll back only a failed file.
        db.session.flush()
    except Exception as e:
        current_app.logger.error(f"Commission ingest error ({carrier}): {e}")
        return {
            "filename": filename, "ok": False,
            "error": f"Could not import {carrier} {period_label}: {e}. No payments were created.",
            "fix": None,
        }

    # Scope label for per-agent carriers (e.g. "Brian Freeman" for BCBS)
    # vs. agency-wide carriers (use carrier name).
    scope_user = User.query.get(agent_id) if agent_id else None
    scope = scope_user.name if scope_user else carrier

    quar = quarantined_line_items(stmt.id, agency_id)
    if quar["count"]:
        warnings.append(
            f"⚠ {quar['count']} {carrier} line(s) totaling ${quar['total']:,.2f} "
            f"need manual review — open the statement's Quarantine tab to split them.")

    return {
        "filename": filename,
        "ok": True,
        "carrier": carrier,
        "scope": scope,
        "rows": ingest.payments_written,
        "gross": stmt.gross_amount,
        "period": period_label,
        "warnings": warnings,
        # Surface ingest detail for the flash message in the route wrapper.
        "_ingest": ingest,
    }


def _process_one_file(filename, file_bytes, statement_month, agency_id, actor,
                      replace=False):
    """Process a single commission file and return a result dict.

    This is the extracted core of the old commission_upload route body.
    It handles BOTH the normalized pipeline (Healthspring/Devoted/BCBS/Aetna/Humana)
    and the legacy CSV/XLSX pipeline (UHC and anything not in NORMALIZERS).

    NEVER raises. NEVER flashes. NEVER redirects. All outcomes are returned
    as a dict.

    TRANSACTION OWNERSHIP: this function calls db.session.flush() — NOT commit().
    The caller (commission_upload route, or the multi-file loop in task 6) owns
    the commit. This allows a db.session.begin_nested() savepoint in a multi-file
    loop to roll back only a single failed file without affecting others.

    Parameters
    ----------
    filename:         original filename string
    file_bytes:       raw uploaded file bytes
    statement_month:  "YYYY-MM" string (from request.form["statement_month"]) or
                      None/empty string for auto-detect
    agency_id:        acting agency id
    actor:            uploading User object (or None in tests); used for
                      uploaded_by_id and agency-scoped lookups
    replace:          if True, bypass the dup-guard and overwrite any existing
                      statement for the same carrier/period

    Returns
    -------
    success: {"filename","ok":True,"carrier","scope","rows","gross","period",
               "warnings":[...], "_ingest":<IngestResult>}
    failure: {"filename","ok":False,"error":<str>,"fix":<str|None>}
    """
    try:
        filename_lower = (filename or "").lower()

        # ── New normalize→resolve→pay pipeline for the 5 clean-split carriers ──
        # These carriers (Healthspring, Devoted, BCBS, Aetna, Humana) ship multi-
        # sheet / SpreadsheetML files that the legacy single-active-sheet path can't
        # read. Detect via sheet_loader (handles xlsx/xls/SpreadsheetML), and if the
        # carrier is in NORMALIZERS, run the ingest pipeline. UHC (and anything not
        # in NORMALIZERS) falls through to the unchanged legacy path below.
        if not filename_lower.endswith(".csv"):
            try:
                _sheets_probe = load_sheets_from_bytes(file_bytes, filename)
            except Exception:
                _sheets_probe = None
            if _sheets_probe:
                probe_carrier = _detect_carrier_from_sheets(_sheets_probe)
                if probe_carrier in NORMALIZERS:
                    return _ingest_normalized_upload(
                        probe_carrier, _sheets_probe, file_bytes, filename,
                        statement_month, agency_id, actor, replace=replace)

        # ── Legacy single-active-sheet pipeline (UHC + CSV) ──
        try:
            if filename_lower.endswith(".csv"):
                wb = _csv_bytes_to_workbook(file_bytes)
            else:
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
        except Exception as e:
            return {"filename": filename, "ok": False,
                    "error": f"Could not read file: {e}", "fix": None}

        carrier = _detect_carrier(ws)
        ok, reason = _carrier_supported_or_reason(carrier)
        if not ok:
            return {"filename": filename, "ok": False, "error": reason, "fix": None}

        try:
            gross, bonus, paid, stmt_date, line_items, stated_rate = PARSERS[carrier](ws)
        except Exception as e:
            current_app.logger.error(f"Commission parse error ({carrier}): {e}")
            return {"filename": filename, "ok": False,
                    "error": f"Parse error for {carrier}: {e}", "fix": None}

        warnings = []

        # 1. Try to get statement month from caller's override
        form_month = (statement_month or "").strip()  # format: "YYYY-MM"
        if form_month:
            try:
                stmt_date = datetime.strptime(form_month, "%Y-%m").date()
            except ValueError:
                pass

        # 2. Fall back to date parsed from file content
        # (already set by parser if it found a date in the file)

        # 3. Try to extract from filename
        if not stmt_date:
            stmt_date = _parse_date_from_filename(filename)

        # 4. Last resort: today (admin will see the period_label and can re-upload with override)
        if not stmt_date:
            stmt_date = date.today()
            warnings.append(
                "Could not detect statement period from file content or filename. "
                "Defaulted to today's month. Use the 'Statement Month' field to correct this.")

        # Every carrier pays Founders (the agency) directly, not any single agent, so
        # the statement is always agency-level: agent_id=None. Per-row PolicyPayment
        # attribution (via build_payments) carries the real per-agent earnings. The
        # statement-level expected/paid math just needs a split rate, taken from any
        # active carrier contract.
        agent_id = None
        contract = AgentCarrierContract.query.filter_by(
            agency_id=agency_id, carrier=carrier, is_active=True
        ).first()
        agent_split = contract.split_rate if contract else 0.55
        period_label = stmt_date.strftime("%B %Y")
        expected     = round((gross + bonus) * agent_split, 2)
        # If the file has no summary row (paid=0), assume expected was paid — no discrepancy.
        # AJ can adjust manually if the actual payment differs.
        if paid == 0.0:
            paid = expected
        difference   = round(expected - paid, 2)
        status       = "verified" if abs(difference) < 0.02 else "discrepancy"

        # Rate discrepancy check — flag when AJ's formula uses a different rate than the contract
        if stated_rate is not None and abs(stated_rate - agent_split) > 0.001:
            stated_pct   = round(stated_rate * 100, 2)
            contract_pct = round(agent_split * 100, 2)
            wrong_expected = round((gross + bonus) * stated_rate, 2)
            rate_diff = round(wrong_expected - expected, 2)
            direction = "underpaid" if rate_diff < 0 else "overpaid"
            warnings.append(
                f"⚠ Rate mismatch on {carrier} {period_label}: AJ's file used {stated_pct}% "
                f"but contract rate is {contract_pct}%. "
                f"This would have {direction} by "
                f"${abs(rate_diff):,.2f}. Portal calculated expected at {contract_pct}%.")

        existing = CommissionStatement.query.filter_by(
            carrier=carrier, agent_id=agent_id, period_label=period_label,
            agency_id=agency_id).first()
        _was_update = existing is not None
        if _was_update:
            warnings.append(
                f"{carrier} {period_label} was already uploaded. "
                "Re-uploading will overwrite the existing statement and payment ledger rows.")
        stmt = existing or CommissionStatement(
            carrier=carrier, agent_id=agent_id, agency_id=agency_id)
        if not existing:
            db.session.add(stmt)

        stmt.statement_date  = stmt_date
        stmt.period_label    = period_label
        stmt.gross_amount    = round(gross + bonus, 2)
        stmt.bonus_amount    = round(bonus, 2)
        stmt.split_rate      = agent_split
        stmt.expected_amount = expected
        stmt.paid_amount     = round(paid, 2)
        stmt.difference      = difference
        stmt.status          = status
        stmt.line_items      = json.dumps(line_items)
        stmt.filename        = filename
        stmt.uploaded_by_id  = actor.id if actor else None
        db.session.flush()   # get stmt.id before building payments

        # If re-uploading, clear stale payment ledger rows for this statement
        if _was_update:
            PolicyPayment.query.filter_by(
                statement_id=stmt.id, agency_id=agency_id
            ).delete(synchronize_session=False)
            db.session.flush()

        # Re-parse worksheet for payment ledger (ws cursor is already at start of data)
        wb2  = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws2  = wb2.active
        build_payments(stmt, carrier, agent_id, agency_id, ws2)
        # NOTE: db.session.flush() here — NOT commit(). Transaction ownership moved
        # to the caller (commission_upload or the multi-file loop in task 6) so that
        # a begin_nested() savepoint can roll back only a failed file.
        db.session.flush()

        split_pct = round(agent_split * 100, 2)
        if status == "verified":
            warnings.insert(0,
                f"✓ {carrier} {period_label} — verified. "
                f"Gross ${stmt.gross_amount:,.2f} × {split_pct}% = ${expected:,.2f} ✅")
        else:
            warnings.insert(0,
                f"⚠ {carrier} {period_label} — discrepancy of ${abs(difference):,.2f}. "
                f"Expected ${expected:,.2f} ({split_pct}%), paid ${round(paid, 2):,.2f}.")

        return {
            "filename": filename,
            "ok": True,
            "carrier": carrier,
            "scope": carrier,          # legacy path is always agency-level
            "rows": len(line_items),
            "gross": stmt.gross_amount,
            "period": period_label,
            "warnings": warnings,
            "_status": status,
            "_expected": expected,
            "_paid": round(paid, 2),
            "_split_pct": split_pct,
            "_difference": difference,
        }

    except Exception as e:
        # Belt-and-suspenders: the inner try/except blocks above catch the known
        # error sites; this catches anything unexpected so the function never raises.
        try:
            current_app.logger.error(f"_process_one_file unexpected error ({filename}): {e}")
        except RuntimeError:
            pass  # no app context in tests — log failure is fine
        return {"filename": filename, "ok": False, "error": str(e), "fix": None}


@commission_bp.route("/admin/commissions/upload", methods=["POST"])
@login_required
def commission_upload():
    """Multi-file commission upload.

    Accepts one or more files via getlist("file"). Reads `statement_month` and
    the `replace` flag ONCE from the form, then processes each file in its own
    savepoint so a failure on one file does not roll back the others.

    XHR / JSON callers (X-Requested-With: XMLHttpRequest or Accept: application/json):
      → 200 JSON {"results": [...], "summary": {"imported": n, "rejected": m}}
      → 400 JSON {"error": "..."} when no files selected

    Non-XHR (back-compat):
      → flash per-file result + redirect to commission_admin
    """
    if not current_user.is_admin:
        abort(403)

    files = [f for f in request.files.getlist("file") if f and f.filename]
    statement_month = request.form.get("statement_month", "").strip()
    replace = request.form.get("replace") == "1"
    is_xhr = (request.headers.get("X-Requested-With") == "XMLHttpRequest"
              or "application/json" in (request.headers.get("Accept") or ""))

    if not files:
        if is_xhr:
            return jsonify(results=[], summary={"imported": 0, "rejected": 0},
                           error="No file selected."), 400
        flash("No file selected.", "error")
        return redirect(url_for("commission.commission_admin"))

    results = []
    for f in files:
        fname = f.filename
        try:
            data = f.read()
        except Exception as e:
            results.append({"filename": fname, "ok": False,
                            "error": f"Could not read upload: {e}", "fix": None})
            continue
        nested = db.session.begin_nested()   # per-file savepoint
        try:
            res = _process_one_file(fname, data, statement_month,
                                    current_user.agency_id, current_user,
                                    replace=replace)
            if res.get("ok"):
                nested.commit()
            else:
                nested.rollback()
            results.append(res)
        except Exception as e:              # defensive — _process_one_file shouldn't raise
            nested.rollback()
            results.append({"filename": fname, "ok": False, "error": str(e), "fix": None})

    db.session.commit()

    summary = {"imported": sum(1 for r in results if r.get("ok")),
               "rejected": sum(1 for r in results if not r.get("ok"))}

    if is_xhr:
        return jsonify(results=results, summary=summary)

    # Non-XHR back-compat: flash per-file result + redirect
    for r in results:
        if r.get("ok"):
            warnings = r.get("warnings", [])
            if r.get("_ingest"):
                ingest = r["_ingest"]
                flash(
                    f"✓ {r['carrier']} {r['period']} — {ingest.payments_written} payments, "
                    f"{ingest.customers_created} customers created "
                    f"({ingest.stubs_created} stubs), {ingest.chargebacks} chargebacks"
                    + (f", {ingest.match_suggestions} match suggestions" if ingest.match_suggestions else "")
                    + ".", "success")
                for w in warnings:
                    flash(w, "warning")
            else:
                if warnings:
                    first = warnings[0]
                    level = "success" if first.startswith("✓") else "warning"
                    flash(first, level)
                    for w in warnings[1:]:
                        flash(w, "warning")
        else:
            flash(f"✗ {r['filename']}: {r['error']}", "error")

    return redirect(url_for("commission.commission_admin"))


@commission_bp.route("/admin/commissions/<int:stmt_id>/delete", methods=["POST"])
@login_required
def commission_delete(stmt_id):
    """Admin deletes a commission statement and all its payment ledger rows."""
    if not current_user.is_admin:
        abort(403)
    stmt = CommissionStatement.query.filter_by(
        id=stmt_id, agency_id=current_user.agency_id).first_or_404()
    label = f"{stmt.carrier} {stmt.period_label}"
    PolicyPayment.query.filter_by(
        statement_id=stmt.id, agency_id=current_user.agency_id
    ).delete(synchronize_session=False)
    db.session.delete(stmt)
    db.session.commit()
    flash(f"{label} statement deleted.", "success")
    return redirect(url_for("commission.commission_admin"))


@commission_bp.route("/admin/commissions/<int:stmt_id>/fidelity")
@login_required
def commission_fidelity(stmt_id):
    """A2 — the Fidelity View: every raw line of a statement beside its agent /
    Founders-override split (G/H), with a footer proving file total = ledger total.
    Lets AJ/Brian confirm the portal reflects the carrier file EXACTLY. Admin-only."""
    if not current_user.is_admin:
        abort(403)
    from app.commission.recap import fidelity_view, balance_status
    stmt = CommissionStatement.query.filter_by(
        id=stmt_id, agency_id=current_user.agency_id).first_or_404()
    fv = fidelity_view(stmt.id, current_user.agency_id)
    bstate, bdelta = balance_status(stmt)
    agents = (User.query.filter_by(agency_id=current_user.agency_id)
              .filter(User.email != "admin@foundersinsuranceagency.com")
              .order_by(User.name).all())
    return render_template("commission_fidelity.html", stmt=stmt, fv=fv,
                           bstate=bstate, bdelta=bdelta, agents=agents)


@commission_bp.route("/admin/commissions/quarantine")
@login_required
def commission_quarantine_workbench():
    """Standalone Quarantine Workbench: every needs_manual_review line across all
    months/carriers/agents (default grouped by month, newest first), with
    period/carrier/agent filters and an amount sort that flattens + clusters
    identical amounts. Admin-only. Resolve/Undo/Edit reuse the existing routes."""
    if not current_user.is_admin:
        abort(403)
    period = request.args.get("period") or None
    carrier = request.args.get("carrier") or None
    agent_id = request.args.get("agent", type=int)
    sort = request.args.get("sort") or None
    if sort not in ("amount_asc", "amount_desc"):
        sort = None
    wb = quarantine_workbench(current_user.agency_id, period=period, carrier=carrier,
                              agent_id=agent_id, sort=sort)
    from app.commission.recap import recently_resolved_workbench
    resolved = recently_resolved_workbench(current_user.agency_id, period=period,
                                           carrier=carrier, agent_id=agent_id)
    agents = (User.query.filter_by(agency_id=current_user.agency_id)
              .filter(User.email != "admin@foundersinsuranceagency.com")
              .order_by(User.name).all())
    return render_template("commission_quarantine_workbench.html", wb=wb, agents=agents,
                           resolved=resolved,
                           f_period=period, f_carrier=carrier, f_agent=agent_id, f_sort=sort)


@commission_bp.route("/admin/commissions/<int:stmt_id>/quarantine")
@login_required
def commission_quarantine(stmt_id):
    """The needs-manual-review lines for a statement (UHC's ~2.3% the parser can't
    auto-split). AJ resolves each in-line: set agent + override $ → the remainder
    splits at the agent's contract rate, the row leaves quarantine."""
    if not current_user.is_admin:
        abort(403)
    stmt = CommissionStatement.query.filter_by(
        id=stmt_id, agency_id=current_user.agency_id).first_or_404()
    quar = quarantined_line_items(stmt.id, current_user.agency_id)
    from app.commission.recap import recently_resolved_line_items
    resolved = recently_resolved_line_items(stmt.id, current_user.agency_id)
    agents = (User.query.filter_by(agency_id=current_user.agency_id)
              .filter(User.email != "admin@foundersinsuranceagency.com")
              .order_by(User.name).all())
    return render_template("commission_quarantine.html", stmt=stmt, quar=quar,
                           resolved=resolved, agents=agents)


@commission_bp.route("/admin/commissions/review")
@login_required
def commission_review():
    """RETIRED — the period-level review page was a duplicate of the Quarantine
    Workbench (same data, confusingly different page). Permanently redirect to the
    Workbench (filtered to the period if one was passed) so there is ONE canonical
    quarantine surface. Kept as a route so old bookmarks/links don't 404."""
    if not current_user.is_admin:
        abort(403)
    period = request.args.get("period")
    return redirect(url_for("commission.commission_quarantine_workbench",
                            period=period) if period
                    else url_for("commission.commission_quarantine_workbench"))


@commission_bp.route("/admin/commissions/line/<int:line_id>/resolve", methods=["POST"])
@login_required
def commission_quarantine_resolve(line_id):
    """Resolve one quarantined line: agent + override $ → agent_commission remainder
    (split at the agent's contract rate) + a founders_override line. Returns to the
    page it was submitted from (per-statement tab or the period review page)."""
    if not current_user.is_admin:
        abort(403)
    from app.commission.ledger import resolve_quarantine_line, NEEDS_MANUAL_REVIEW
    li = CommissionLineItem.query.filter_by(
        id=line_id, agency_id=current_user.agency_id).first_or_404()

    # Return to wherever the resolve was triggered from.
    back = (request.form.get("next") or request.referrer
            or url_for("commission.commission_quarantine", stmt_id=li.statement_id))

    if li.classification != NEEDS_MANUAL_REVIEW:
        flash("That line was already resolved.", "warning")
        return redirect(back)

    agent = User.query.filter_by(id=request.form.get("agent_id", type=int),
                                 agency_id=current_user.agency_id).first()
    if not agent:
        flash("Pick a valid agent.", "error")
        return redirect(back)
    try:
        override_amount = float(request.form.get("override_amount") or 0)
    except ValueError:
        override_amount = 0.0

    # split rate from the agent's contract for this carrier (fallback 0.55).
    contract = AgentCarrierContract.query.filter_by(
        agent_id=agent.id, carrier=li.carrier, is_active=True,
        agency_id=current_user.agency_id).first()
    split_rate = contract.split_rate if contract else 0.55

    try:
        resolve_quarantine_line(li, agent.id, override_amount, split_rate,
                                user_id=current_user.id)
        db.session.commit()
    except ValueError as e:
        db.session.rollback()
        flash(f"Could not resolve: {e}", "error")
        return redirect(back)
    from app.audit import log_event
    log_event("commission_resolve", category="commission",
              detail=f"{li.carrier} {li.member_name or 'line'} -> agent {agent.id} "
                     f"override ${override_amount:.2f}")
    flash(f"Resolved {li.member_name or 'line'} → {agent.display_name} "
          f"(split {split_rate:.0%}, override ${override_amount:,.2f}).", "success")
    return redirect(back)


@commission_bp.route("/admin/commissions/line/<int:line_id>/undo", methods=["POST"])
@login_required
def commission_line_undo(line_id):
    """Undo the most recent human change to a commission line (admin-only)."""
    if not current_user.is_admin:
        abort(403)
    from app.commission.ledger import undo_last_change
    from app.audit import log_event
    li = CommissionLineItem.query.filter_by(
        id=line_id, agency_id=current_user.agency_id).first_or_404()
    back = (request.form.get("next") or request.referrer
            or url_for("commission.commission_quarantine", stmt_id=li.statement_id))
    if undo_last_change(li, user_id=current_user.id):
        db.session.commit()
        log_event("commission_undo", category="commission",
                  detail=f"{li.carrier} {li.member_name or 'line'} #{li.id}")
        flash("Change undone.", "success")
    else:
        flash("Nothing to undo on that line.", "warning")
    return redirect(back)


@commission_bp.route("/admin/commissions/line/<int:line_id>/edit", methods=["POST"])
@login_required
def commission_line_edit(line_id):
    """Correct a line's agent/override split (admin-only, invariant-safe). The
    agent's split_rate is ALWAYS looked up from their real AgentCarrierContract for
    this carrier (never hardcoded) — different agents have different rates (e.g.
    Betty Marlowe = 52.5%, not 55%), so a wrong constant would silently corrupt pay."""
    if not current_user.is_admin:
        abort(403)
    from app.commission.ledger import edit_line_split
    from app.commission.recap import fidelity_row, fidelity_view
    from app.audit import log_event
    # AJAX (Fidelity view) wants JSON to repaint one row in place — no page reload;
    # a plain form POST (no-JS fallback) still redirects back. Detect either signal.
    wants_json = (request.headers.get("X-Requested-With") == "XMLHttpRequest"
                  or "application/json" in (request.headers.get("Accept") or ""))

    def _fail(msg, code=400):
        if wants_json:
            return jsonify(ok=False, error=msg), code
        flash(msg, "error")
        return redirect(back)

    li = CommissionLineItem.query.filter_by(
        id=line_id, agency_id=current_user.agency_id).first_or_404()
    back = (request.form.get("next") or request.referrer
            or url_for("commission.commission_quarantine", stmt_id=li.statement_id))
    agent = User.query.filter_by(id=request.form.get("agent_id", type=int),
                                 agency_id=current_user.agency_id).first()
    if not agent:
        return _fail("Pick a valid agent.")
    try:
        agent_amount = float(request.form.get("agent_amount") or 0)
        override_amount = float(request.form.get("override_amount") or 0)
    except ValueError:
        return _fail("Enter valid amounts.")

    # AJ enters the EXACT dollars (agent + Founders override). edit_line_split stores
    # the agent amount as the final payout (split_rate=1.0), so no contract rate is
    # re-applied — this is what lets a special case (e.g. Anjana keeps 100% of the
    # post-override amount) flow through to her recap unchanged.
    try:
        edit_line_split(li, agent_amount=agent_amount, override_amount=override_amount,
                        agent_id=agent.id, user_id=current_user.id)
        db.session.commit()
    except ValueError as e:
        db.session.rollback()
        return _fail(f"Could not edit: {e}")
    log_event("commission_edit", category="commission",
              detail=f"{li.carrier} {li.member_name or 'line'} #{li.id} "
                     f"-> agent ${agent_amount:.2f} / override ${override_amount:.2f}")

    if wants_json:
        # Rebuild the edited row (and its ::ovr sibling, if any) from the SAME
        # fidelity_row logic the table uses, plus refreshed statement totals.
        db.session.refresh(li)
        ovr = (CommissionLineItem.query
               .filter_by(statement_id=li.statement_id,
                          source_ref=f"{li.source_ref}::ovr").first())
        fv = fidelity_view(li.statement_id, current_user.agency_id)
        return jsonify(
            ok=True,
            row=fidelity_row(li),
            sibling=fidelity_row(ovr) if ovr is not None else None,
            totals={"raw_total": fv["raw_total"], "agent_total": fv["agent_total"],
                    "founders_total": fv["founders_total"]})

    flash("Split updated.", "success")
    return redirect(back)


@commission_bp.route("/admin/commissions/agent/<int:agent_id>")
@login_required
def commission_agent_detail(agent_id):
    if not current_user.is_admin:
        abort(403)
    agent = User.query.get_or_404(agent_id)
    statements = (CommissionStatement.query
                  .filter_by(agent_id=agent_id, agency_id=current_user.agency_id)
                  .order_by(CommissionStatement.statement_date.desc())
                  .all())
    for s in statements:
        s.line_items_parsed = json.loads(s.line_items) if s.line_items else []
    _enrich_line_items_with_commission_type(statements, current_user.agency_id)
    return render_template("commission.html",
        statements=statements, is_admin=True, viewing_agent=agent)


# ── Payment ledger ────────────────────────────────────────────────────────────

@commission_bp.route("/commissions/ledger")
@login_required
def commission_ledger():
    """Per-member payment ledger — agent view."""
    from sqlalchemy import func as sqlfunc

    agency_id = current_user.agency_id
    agent_id  = current_user.id

    # Available periods for this agent
    periods = (db.session.query(PolicyPayment.period_label, PolicyPayment.statement_date)
               .filter_by(agent_id=agent_id, agency_id=agency_id)
               .distinct()
               .order_by(PolicyPayment.statement_date.desc())
               .all())

    selected_period = request.args.get("period") or (periods[0].period_label if periods else None)
    carrier_filter  = request.args.get("carrier", "all")
    action_filter   = request.args.get("action",  "all")

    payments = []
    carriers = []
    summary  = {}

    if selected_period:
        q = (PolicyPayment.query
             .filter_by(agent_id=agent_id, agency_id=agency_id,
                        period_label=selected_period))
        if carrier_filter != "all":
            q = q.filter_by(carrier=carrier_filter)
        if action_filter != "all":
            q = q.filter_by(commission_action=action_filter)

        payments = q.order_by(PolicyPayment.carrier, PolicyPayment.member_name).all()

        # Summary stats for selected period (all carriers, unfiltered)
        all_period = (PolicyPayment.query
                      .filter_by(agent_id=agent_id, agency_id=agency_id,
                                 period_label=selected_period)
                      .all())
        total_paid      = sum(p.paid_amount for p in all_period)
        total_members   = len([p for p in all_period if not p.is_chargeback])
        total_chargebacks = sum(p.paid_amount for p in all_period if p.is_chargeback)
        unmatched_count = sum(1 for p in all_period if p.match_confidence == "unmatched")
        carriers = sorted(set(p.carrier for p in all_period))

        summary = {
            "total_paid":        total_paid,
            "total_members":     total_members,
            "total_chargebacks": total_chargebacks,
            "unmatched_count":   unmatched_count,
            "net_paid":          total_paid + total_chargebacks,
        }

    return render_template("commission_ledger.html",
        periods=periods,
        selected_period=selected_period,
        carrier_filter=carrier_filter,
        action_filter=action_filter,
        payments=payments,
        carriers=carriers,
        summary=summary,
        is_admin=False,
        viewing_agent=None,
    )


@commission_bp.route("/admin/commissions/ledger")
@login_required
def commission_ledger_admin():
    """Per-member payment ledger — admin view, all agents."""
    if not current_user.is_admin:
        abort(403)

    agency_id    = current_user.agency_id
    agent_id_arg = request.args.get("agent_id", type=int)

    agents = (User.query
              .filter(User.email != "admin@foundersinsuranceagency.com",
                      User.agency_id == agency_id)
              .order_by(User.name).all())

    selected_agent_id = agent_id_arg or (agents[0].id if agents else None)

    stale_parked_count = parked_payments_older_than(30, agency_id)

    periods = []
    if selected_agent_id:
        periods = (db.session.query(PolicyPayment.period_label, PolicyPayment.statement_date)
                   .filter_by(agent_id=selected_agent_id, agency_id=agency_id)
                   .distinct()
                   .order_by(PolicyPayment.statement_date.desc())
                   .all())

    selected_period = request.args.get("period") or (periods[0].period_label if periods else None)
    carrier_filter  = request.args.get("carrier", "all")
    action_filter   = request.args.get("action",  "all")

    payments = []
    carriers = []
    summary  = {}

    if selected_period and selected_agent_id:
        q = (PolicyPayment.query
             .filter_by(agent_id=selected_agent_id, agency_id=agency_id,
                        period_label=selected_period))
        if carrier_filter != "all":
            q = q.filter_by(carrier=carrier_filter)
        if action_filter != "all":
            q = q.filter_by(commission_action=action_filter)

        payments = q.order_by(PolicyPayment.carrier, PolicyPayment.member_name).all()

        all_period = (PolicyPayment.query
                      .filter_by(agent_id=selected_agent_id, agency_id=agency_id,
                                 period_label=selected_period)
                      .all())
        total_paid        = sum(p.paid_amount for p in all_period)
        total_members     = len([p for p in all_period if not p.is_chargeback])
        total_chargebacks = sum(p.paid_amount for p in all_period if p.is_chargeback)
        unmatched_count   = sum(1 for p in all_period if p.match_confidence == "unmatched")
        carriers = sorted(set(p.carrier for p in all_period))

        summary = {
            "total_paid":        total_paid,
            "total_members":     total_members,
            "total_chargebacks": total_chargebacks,
            "unmatched_count":   unmatched_count,
            "net_paid":          total_paid + total_chargebacks,
        }

    return render_template("commission_ledger.html",
        periods=periods,
        selected_period=selected_period,
        carrier_filter=carrier_filter,
        action_filter=action_filter,
        payments=payments,
        carriers=carriers,
        summary=summary,
        is_admin=True,
        viewing_agent=User.query.get(selected_agent_id) if selected_agent_id else None,
        agents=agents,
        selected_agent_id=selected_agent_id,
        stale_parked_count=stale_parked_count,
    )


# ── Override workflow ──────────────────────────────────────────────────────────

@commission_bp.route("/admin/commissions/<int:stmt_id>/request-override", methods=["POST"])
@login_required
def commission_request_override(stmt_id):
    """Admin submits an explanation for a discrepancy and sends it to the agent for review."""
    if not current_user.is_admin:
        abort(403)
    stmt = CommissionStatement.query.filter_by(
        id=stmt_id, agency_id=current_user.agency_id).first_or_404()
    if stmt.status not in ("discrepancy",):
        flash("Override can only be requested on statements with a discrepancy.", "error")
        return redirect(url_for("commission.commission_admin"))

    note = request.form.get("override_note_admin", "").strip()
    if not note:
        flash("An explanation is required to submit for agent review.", "error")
        return redirect(url_for("commission.commission_admin"))

    stmt.override_note_admin     = note
    stmt.override_requested_by_id = current_user.id
    stmt.override_requested_at   = datetime.utcnow()
    stmt.override_note_agent     = None
    stmt.override_reviewed_by_id = None
    stmt.override_reviewed_at    = None
    stmt.status                  = "pending_review"
    db.session.commit()

    agent = User.query.get(stmt.agent_id)
    flash(f"Override sent to {agent.display_name} for review on {stmt.carrier} {stmt.period_label}.", "success")
    return redirect(url_for("commission.commission_admin"))


@commission_bp.route("/commissions/<int:stmt_id>/review-override", methods=["POST"])
@login_required
def commission_review_override(stmt_id):
    """Agent accepts or disputes an override submitted by admin."""
    stmt = CommissionStatement.query.filter_by(
        id=stmt_id, agent_id=current_user.id,
        agency_id=current_user.agency_id).first_or_404()
    if stmt.status != "pending_review":
        flash("This statement is not awaiting your review.", "error")
        return redirect(url_for("commission.commission_index"))

    action = request.form.get("action")   # "accept" or "dispute"
    note   = request.form.get("override_note_agent", "").strip()

    if action not in ("accept", "dispute"):
        flash("Invalid action.", "error")
        return redirect(url_for("commission.commission_index"))
    if action == "dispute" and not note:
        flash("Please provide your reasoning when disputing a discrepancy.", "error")
        return redirect(url_for("commission.commission_index"))

    stmt.override_note_agent     = note
    stmt.override_reviewed_by_id = current_user.id
    stmt.override_reviewed_at    = datetime.utcnow()
    stmt.status                  = "accepted" if action == "accept" else "disputed"
    db.session.commit()

    if action == "accept":
        flash(f"You accepted the {stmt.carrier} {stmt.period_label} override.", "success")
    else:
        flash(f"Your dispute on {stmt.carrier} {stmt.period_label} has been submitted to admin for review.", "warning")
    return redirect(url_for("commission.commission_index"))


@commission_bp.route("/admin/commissions/<int:stmt_id>/close-dispute", methods=["POST"])
@login_required
def commission_close_dispute(stmt_id):
    """Admin closes a disputed statement — marks it accepted after reviewing agent's note."""
    if not current_user.is_admin:
        abort(403)
    stmt = CommissionStatement.query.filter_by(
        id=stmt_id, agency_id=current_user.agency_id).first_or_404()
    if stmt.status != "disputed":
        flash("This statement is not in disputed status.", "error")
        return redirect(url_for("commission.commission_admin"))

    stmt.status = "accepted"
    db.session.commit()
    flash(f"{stmt.carrier} {stmt.period_label} dispute closed and marked accepted.", "success")
    return redirect(url_for("commission.commission_admin"))


# ─────────────────────────────────────────────
# Reconciliation helpers
# ─────────────────────────────────────────────

def _period_bounds(period_label):
    """Convert 'March 2026' to (date(2026,3,1), date(2026,3,31))."""
    try:
        start = datetime.strptime(period_label, "%B %Y").date().replace(day=1)
        end = start + relativedelta(months=1) - relativedelta(days=1)
        return start, end
    except (ValueError, TypeError):
        return None, None


def _reconcile(agency_id, agent_id, carrier, period_label):
    """Run both reconciliation queries for a single agent+carrier+period.
    Returns dict with 'unpaid_policies' and 'unmatched_payments'.
    """
    period_start, period_end = _period_bounds(period_label)
    if not period_start:
        return {'unpaid_policies': [], 'unmatched_payments': []}

    paid_policy_ids = (db.session.query(PolicyPayment.policy_id)
        .filter_by(agency_id=agency_id, agent_id=agent_id,
                   carrier=carrier, period_label=period_label)
        .filter(PolicyPayment.policy_id.isnot(None))
        .subquery())

    unpaid = (Policy.query
        .filter_by(agency_id=agency_id, agent_id=agent_id,
                   carrier=carrier, status='active')
        .filter(Policy.effective_date <= period_end)
        .filter(or_(Policy.term_date.is_(None), Policy.term_date > period_start))
        .filter(~Policy.id.in_(paid_policy_ids))
        .all())

    unmatched = (PolicyPayment.query
        .filter_by(agency_id=agency_id, agent_id=agent_id,
                   carrier=carrier, period_label=period_label,
                   match_confidence='unmatched')
        .all())

    return {'unpaid_policies': unpaid, 'unmatched_payments': unmatched}


# ─────────────────────────────────────────────
# Reconciliation routes
# ─────────────────────────────────────────────

@commission_bp.route('/commissions/reconciliation')
@login_required
def reconciliation_view():
    agency_id = current_user.agency_id
    agent_id = current_user.id

    available = (db.session.query(
            CommissionStatement.carrier,
            CommissionStatement.period_label,
        )
        .filter_by(agency_id=agency_id, agent_id=agent_id)
        .distinct()
        .order_by(CommissionStatement.period_label.desc(),
                  CommissionStatement.carrier.asc())
        .all())

    selected_carrier = request.args.get('carrier')
    selected_period = request.args.get('period')

    results = None
    if selected_carrier and selected_period:
        results = _reconcile(agency_id, agent_id, selected_carrier, selected_period)

    return render_template('commission_reconciliation.html',
        available=available,
        selected_carrier=selected_carrier,
        selected_period=selected_period,
        results=results,
        is_admin=False,
        agents=None,
        selected_agent_id=agent_id)


@commission_bp.route('/admin/commissions/reconciliation')
@login_required
def admin_reconciliation_view():
    if not current_user.is_admin:
        abort(403)
    agency_id = current_user.agency_id

    agents = User.query.filter_by(agency_id=agency_id).order_by(User.email).all()

    selected_agent_id = request.args.get('agent_id', type=int) or current_user.id

    available = (db.session.query(
            CommissionStatement.carrier,
            CommissionStatement.period_label,
        )
        .filter_by(agency_id=agency_id, agent_id=selected_agent_id)
        .distinct()
        .order_by(CommissionStatement.period_label.desc(),
                  CommissionStatement.carrier.asc())
        .all())

    selected_carrier = request.args.get('carrier')
    selected_period = request.args.get('period')

    results = None
    if selected_carrier and selected_period:
        results = _reconcile(agency_id, selected_agent_id, selected_carrier, selected_period)

    return render_template('commission_reconciliation.html',
        available=available,
        selected_carrier=selected_carrier,
        selected_period=selected_period,
        results=results,
        is_admin=True,
        agents=agents,
        selected_agent_id=selected_agent_id)


# ---------------------------------------------------------------------------
# R2 — Agent Commission Recap routes
# ---------------------------------------------------------------------------

def _published_periods(agent_id, agency_id):
    rows = (AgentRecapPeriod.query
            .filter_by(agency_id=agency_id, agent_id=agent_id, status="published")
            .order_by(AgentRecapPeriod.published_at.desc()).all())
    return [r.period_label for r in rows]


def _latest_published_period(agent_id, agency_id):
    ps = _published_periods(agent_id, agency_id)
    return ps[0] if ps else None


@commission_bp.route("/commissions/recap")
@login_required
def agent_recap():
    period = request.args.get("period") or _latest_published_period(current_user.id, current_user.agency_id)
    rp = AgentRecapPeriod.query.filter_by(
        agency_id=current_user.agency_id, agent_id=current_user.id, period_label=period).first() if period else None
    if not period or not is_visible_to_agent(rp):
        return render_template("commission/recap.html", recap=None, pending=True,
                               period_label=period, is_admin=current_user.is_admin,
                               periods=_published_periods(current_user.id, current_user.agency_id))
    recap = build_recap(current_user.id, current_user.agency_id, period)
    return render_template("commission/recap.html", recap=recap, pending=False,
                           period_label=period, is_admin=current_user.is_admin,
                           periods=_published_periods(current_user.id, current_user.agency_id))


@commission_bp.route("/admin/commissions/aggregate")
@login_required
def admin_aggregate():
    """All agents × all carriers commission matrix (Option A). Payout + Founders-
    keep per cell, month / year-to-date toggle, click a cell to drill into that
    agent+carrier's recap."""
    if not current_user.is_admin:
        abort(403)
    scope = "ytd" if request.args.get("scope") == "ytd" else "month"
    period = (request.args.get("period")
              or latest_period_with_data(current_user.agency_id)
              or date.today().strftime("%B %Y"))
    try:
        year = datetime.strptime(period, "%B %Y").year
    except ValueError:
        year = date.today().year
    matrix = build_aggregate_matrix(current_user.agency_id, scope=scope,
                                    period_label=period, year=year)
    # Per-carrier data status for the period (received / confirmed_zero / pending),
    # so a column header shows whether that carrier's statement is in yet.
    from app.commission.recap import carrier_period_status
    statuses = ({c: carrier_period_status(current_user.agency_id, period, c)
                 for c in matrix["carriers"]} if scope == "month" else {})
    # Persistent agent nav bar (same 1-click pills as the recap page); admin@ excluded.
    agents = (User.query.filter_by(agency_id=current_user.agency_id)
              .filter(User.email != "admin@foundersinsuranceagency.com")
              .order_by(User.name).all())
    return render_template("commission/aggregate.html", matrix=matrix, scope=scope,
                           period_label=period, year=year, carrier_status=statuses,
                           agents=agents,
                           periods=all_periods_with_data(current_user.agency_id),
                           is_admin=True)


@commission_bp.route("/admin/commissions/confirm-zero", methods=["POST"])
@login_required
def admin_confirm_zero():
    """AJ confirms a carrier had NO business for a period (genuine $0), so a blank
    cell reads 'confirmed $0' instead of 'statement not uploaded yet'."""
    if not current_user.is_admin:
        abort(403)
    from app.models import CarrierPeriodConfirmation, CommissionStatement
    carrier = (request.form.get("carrier") or "").strip()
    period = (request.form.get("period") or "").strip()
    note = (request.form.get("note") or "").strip() or None
    if not (carrier and period):
        flash("Carrier and period required.", "error")
        return redirect(url_for("commission.admin_aggregate", period=period))
    # If a statement already exists for this carrier+period it's already 'received';
    # confirming zero is only meaningful when there's no statement.
    existing = CarrierPeriodConfirmation.query.filter_by(
        agency_id=current_user.agency_id, carrier=carrier, period_label=period).first()
    if not existing:
        db.session.add(CarrierPeriodConfirmation(
            agency_id=current_user.agency_id, carrier=carrier, period_label=period,
            confirmed_by_id=current_user.id, note=note))
        db.session.commit()
    flash(f"Confirmed $0 for {carrier} — {period}.", "success")
    return redirect(url_for("commission.admin_aggregate", period=period))


@commission_bp.route("/admin/commissions/recap")
@login_required
def admin_recap():
    if not current_user.is_admin:
        abort(403)
    # The admin commission landing IS the All-Commissions matrix (the agency
    # overview). A specific agent's recap is reached by clicking a matrix row/cell
    # (which passes agent_id). Without agent_id, don't show AJ's own empty "My
    # Commission" shell — send him to the matrix.
    agent_id = request.args.get("agent_id", type=int)
    if agent_id is None:
        return redirect(url_for("commission.admin_aggregate",
                                period=request.args.get("period")))
    # Default to the most recent period that actually has commission data (so opening
    # the page after an upload lands on that period), falling back to today's month.
    period = (request.args.get("period")
              or latest_period_with_data(current_user.agency_id)
              or date.today().strftime("%B %Y"))
    rp = get_or_create_period(agent_id, current_user.agency_id, period)
    db.session.commit()
    recap = build_recap(agent_id, current_user.agency_id, period)
    # Agent nav bar (#4): real agents only — exclude the shared admin@ account (it's
    # never a commission agent; same filter as the Commission Audit view).
    agents = (User.query.filter_by(agency_id=current_user.agency_id)
              .filter(User.email != "admin@foundersinsuranceagency.com")
              .order_by(User.name).all())
    # Period-level quarantine surfacing: any statement this period with needs-review
    # lines (UHC's ~2.3%) so AJ sees them from where he reviews commissions.
    quar_links = []
    stmts = CommissionStatement.query.filter_by(
        agency_id=current_user.agency_id, period_label=period).all()
    for s in stmts:
        q = quarantined_line_items(s.id, current_user.agency_id)
        if q["count"]:
            quar_links.append({"stmt_id": s.id, "carrier": s.carrier,
                               "count": q["count"], "total": q["total"]})
    from app.models import CommissionAdjustment
    adjustments = (CommissionAdjustment.query
                   .filter_by(agency_id=current_user.agency_id, agent_id=agent_id,
                              period_label=period)
                   .order_by(CommissionAdjustment.carrier).all())
    carriers_for_adj = [b.carrier for b in recap.carriers] if recap else []
    return render_template("commission/recap.html", recap=recap, pending=False, admin_view=True,
                           period_label=period, recap_period=rp, agents=agents,
                           quar_links=quar_links, adjustments=adjustments,
                           carriers_for_adj=carriers_for_adj,
                           periods=all_periods_with_data(current_user.agency_id),
                           selected_agent_id=agent_id, is_admin=True)


@commission_bp.route("/admin/commissions/recap/publish", methods=["POST"])
@login_required
def admin_recap_publish():
    if not current_user.is_admin:
        abort(403)
    agent_id = request.form.get("agent_id", type=int)
    period = request.form.get("period")
    rp = get_or_create_period(agent_id, current_user.agency_id, period)
    recap = build_recap(agent_id, current_user.agency_id, period)
    agent = User.query.get(agent_id)
    publish_recap(rp, published_by_id=current_user.id,
                  agent_email=(agent.email if agent else None),
                  total_paid=recap.total_paid, base_url=request.url_root.rstrip("/"))
    db.session.commit()
    flash(f"Published {period} recap for {agent.name if agent else agent_id}.", "success")
    return redirect(url_for("commission.admin_recap", agent_id=agent_id, period=period))


@commission_bp.route("/admin/commissions/recap/set-uhc", methods=["POST"])
@login_required
def admin_recap_set_uhc():
    if not current_user.is_admin:
        abort(403)
    agent_id = request.form.get("agent_id", type=int)
    period = request.form.get("period")
    rp = get_or_create_period(agent_id, current_user.agency_id, period)
    raw = (request.form.get("uhc_amount") or "").replace("$", "").replace(",", "").strip()
    rp.uhc_manual_amount = float(raw) if raw else None
    rp.uhc_manual_note = (request.form.get("uhc_note") or "").strip() or None
    db.session.commit()
    flash("UHC figure updated.", "success")
    return redirect(url_for("commission.admin_recap", agent_id=agent_id, period=period))


@commission_bp.route("/admin/commissions/recap/adjustment", methods=["POST"])
@login_required
def admin_recap_add_adjustment():
    """AJ adds a manual reconciliation line to an agent's carrier block for a period."""
    if not current_user.is_admin:
        abort(403)
    from app.models import CommissionAdjustment
    agent_id = request.form.get("agent_id", type=int)
    period = request.form.get("period")
    carrier = (request.form.get("carrier") or "").strip()
    note = (request.form.get("note") or "").strip()
    raw = (request.form.get("amount") or "").replace("$", "").replace(",", "").strip()
    try:
        amount = float(raw)
    except (ValueError, TypeError):
        amount = None
    if not (agent_id and period and carrier and note and amount is not None):
        flash("Adjustment needs an agent, carrier, amount, and note.", "error")
        return redirect(url_for("commission.admin_recap", agent_id=agent_id, period=period))
    db.session.add(CommissionAdjustment(
        agency_id=current_user.agency_id, agent_id=agent_id, carrier=carrier,
        period_label=period, amount=amount, note=note, created_by_id=current_user.id))
    db.session.commit()
    flash(f"Adjustment added to {carrier} {period}: ${amount:,.2f}.", "success")
    return redirect(url_for("commission.admin_recap", agent_id=agent_id, period=period))


@commission_bp.route("/admin/commissions/recap/adjustment/<int:adj_id>/delete", methods=["POST"])
@login_required
def admin_recap_delete_adjustment(adj_id):
    if not current_user.is_admin:
        abort(403)
    from app.models import CommissionAdjustment
    adj = CommissionAdjustment.query.filter_by(
        id=adj_id, agency_id=current_user.agency_id).first_or_404()
    agent_id, period = adj.agent_id, adj.period_label
    db.session.delete(adj)
    db.session.commit()
    flash("Adjustment removed.", "success")
    return redirect(url_for("commission.admin_recap", agent_id=agent_id, period=period))


@commission_bp.route("/commissions/recap/carrier")
@login_required
def recap_carrier_detail():
    """JSON: the grouped line items for one carrier (lazy-loaded drill-down + search)."""
    agent_id = request.args.get("agent_id", type=int) or current_user.id
    if agent_id != current_user.id and not current_user.is_admin:
        abort(403)
    period = request.args.get("period")
    carrier = request.args.get("carrier")
    q = (request.args.get("q") or "").strip().lower()
    from app.commission.recap import canon_carrier
    blocks = build_carrier_blocks(agent_id, current_user.agency_id, period)
    # Match by CANONICAL name so clicking the "Wellabe" chip finds the underlying
    # Medico/Wellable block(s) — the recap page collapses those to one Wellabe chip.
    matching = [b for b in blocks if canon_carrier(b.carrier) == carrier]
    if not matching:
        return {"carrier": carrier, "groups": []}

    def rowj(r):
        # payout is already rounded to cents at the source (build_carrier_blocks);
        # round raw for display too so the columns shown reconcile to the totals.
        return {"member": r.member_name, "customer_id": r.customer_id, "type": r.type_label,
                "kind": r.type_kind, "raw": round(r.raw_amount or 0.0, 2),
                "split": r.split_rate, "payout": r.payout}

    groups = []
    total = 0.0
    for block in matching:
        total += block.total_payout
        for g in block.groups:
            rows = [rowj(r) for r in g.rows if not q or q in (r.member_name or "").lower()]
            groups.append({"kind": g.kind, "count": g.count, "subtotal": g.subtotal, "rows": rows})
    return {"carrier": carrier, "total": total, "groups": groups}
