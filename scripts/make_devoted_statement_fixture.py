"""
scripts/make_devoted_statement_fixture.py

One-shot generator for tests/fixtures/commission/devoted_statement_sample.xlsx —
a sanitized copy of Devoted's per-agent STATEMENT format (Summary/Detail/Misc),
matching the real 20182775_Rebekah_Long file's column layout and totals.

Run: python3 scripts/make_devoted_statement_fixture.py
"""
import os
import openpyxl

OUT = os.path.join(os.path.dirname(__file__), "..", "tests", "fixtures",
                   "commission", "devoted_statement_sample.xlsx")

DETAIL_HEADER = ["Statement Date", "Agent NPN", "Agent Name", "Member ID",
                 "Member HICN", "Member First", "Member Last", "Member State",
                 "Signature Date", "Effective Date", "Disenroll Date", "Contract",
                 "PBP", "Prior Plan Type", "CMS Cycle Year", "Commission Type",
                 "Period", "Base Amount", "Admin Amount", "Total Payment", "FMO",
                 "Payment Notations"]


def _detail_row(member_id, hicn, first, last, base):
    return ["05/29/2026", "20182775", "Rebekah Long", member_id, hicn, first, last,
            "NC", "11/06/2025", "12/01/2025", "", "H9700", "2", "NONE", "2",
            "Renewal - Monthly", "May", base, 0, base, "Tidewater Management", ""]


def main():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Summary"
    for row in [
        ["Description", "Member Count", "Total", "", "Payee", "Payment Date"],
        ["Credits", 2, "$57.82", "", "Rebekah Long", "05/29/2026"],
        ["Debits", 0, "$0.00", "", "", ""],
        ["Miscellaneous", 8, "($400.00)", "", "", ""],
        ["Bonus", "", "$0.00", "", "", ""],
        ["Sub Total", 10, "($718.11)", "", "", ""],
        ["Balance", "", "($375.93)", "", "", ""],
        ["TOTAL", "", "($718.11)", "", "", ""],
    ]:
        ws.append(row)

    det = wb.create_sheet("Detail")
    det.append(DETAIL_HEADER)
    det.append(_detail_row("DAH887", "7QY9GM5CA40", "MICHELLE", "BROADWAY", 28.91))
    det.append(_detail_row("DAUU67", "6VT3RT2FM11", "BOBBY", "SMITH", 28.91))

    misc = wb.create_sheet("Misc")
    misc.append(["Rep Name", "Rep ID", "Amount", "Note"])
    for note in ["Debra", "James", "Sarah", "DAVID", "DONNA", "Mark", "RITA", "Charlie"]:
        misc.append(["Rebekah Long", "20182775", "($50.00)", f"HRA for member {note}"])

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("wrote", os.path.abspath(OUT))


if __name__ == "__main__":
    main()
