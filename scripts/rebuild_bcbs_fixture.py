"""Regenerate tests/fixtures/commission/bcbs_sample.xlsx to the REAL 14-column
BCBS layout by dropping the phantom 'Customer Type' column (index 3) from every
row. Real BCBS/Tidewater files have NO Customer Type column; the old fixture had
one, which is why the parser (built against it) matched a layout that never existed
and found 0 rows in real files. Idempotent-ish: only drops a col named exactly
'Customer Type'. Run: python3 scripts/rebuild_bcbs_fixture.py
"""
import openpyxl

FIX = "tests/fixtures/commission/bcbs_sample.xlsx"


def main():
    wb = openpyxl.load_workbook(FIX)
    ws = wb["Sheet1"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    header = [str(h or "").strip() for h in rows[0]]
    if "Customer Type" not in header:
        print("Fixture already has no 'Customer Type' column — nothing to do.")
        return
    drop = header.index("Customer Type")
    new_rows = [[c for i, c in enumerate(r) if i != drop] for r in rows]
    # rewrite the sheet
    ws.delete_rows(1, ws.max_row)
    for r in new_rows:
        ws.append(r)
    wb.save(FIX)
    print(f"Dropped 'Customer Type' (col {drop}); fixture now "
          f"{len(new_rows[0])} cols, header[3]={new_rows[0][3]!r}, "
          f"header[4]={new_rows[0][4]!r}, header[-1]={new_rows[0][-1]!r}")


if __name__ == "__main__":
    main()
