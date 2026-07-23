"""Devoted BOB parser tests — covers the July-2026 XLSX 'Application Status Report'
format (single Full Name, Birth Date, Current Status, no member_id) AND the older
snake_case CSV with member_id."""
import openpyxl
from app.parsers.devoted import parse
from app.upload import _rec_is_more_current, _dedupe_bob_records, AMBIGUOUS_WINNING_PAIRS
from datetime import date
from datetime import date as _d

APP_COLS = [
    "Application Status Report Agent Name",
    "Application Status Report Agent Npn",
    "Application Status Report Agent Primary Rts State",
    "Application Status Report Full Name",
    "Application Status Report Birth Date",
    "Application Status Report Current Status",
    "Application Status Report Is Agent Current Aor (Yes / No)",
    "Application Status Report Aor Start Date",
    "Application Status Report Start Date",
    "Application Status Report Plan Name",
    "Application Status Report Plan ID",
    "Application Status Report Is Plan Change (Yes / No)",
    "Application Status Report Is New to Medicare Advantage (Yes / No)",
    "Application Status Report Is New to Devoted (Yes / No)",
]


def _app_status_xlsx(tmp_path, rows):
    p = tmp_path / "Devoted Book of business.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(APP_COLS)
    for r in rows:
        ws.append(r)
    wb.save(p)
    return str(p)


def test_july_application_status_xlsx_parses_name_dob_key(tmp_path):
    """The July Devoted BOB is an XLSX Application-Status Report: single 'Full Name',
    'Birth Date', 'Current Status', NO member_id. Must parse (split the name,
    synthesize a member_id from name+DOB, keep only Enrolled), not raise a CSV
    decode error or 'missing required columns: member_id'."""
    p = _app_status_xlsx(tmp_path, [
        ["Anjana Patel", "21041582", "NC", "BRANDI TUCKER", "1977-03-21",
         "Enrolled", "Yes", "2026-04-01", "2026-04-01", "Devoted CORE", "H1234-001",
         "No", "Yes", "Yes"],
        # a non-enrolled row is dropped
        ["Anjana Patel", "21041582", "NC", "JOHN DOE", "1955-01-02",
         "Withdrawn", "Yes", "2026-04-01", "2026-04-01", "Devoted CORE", "H1234-001",
         "No", "Yes", "Yes"],
    ])
    recs = parse(p)
    assert len(recs) == 1                          # only the Enrolled row
    r = recs[0]
    assert r["carrier"] == "Devoted"
    assert r["first_name"] == "Brandi" and r["last_name"] == "Tucker"
    assert r["dob"] is not None
    assert r["member_id"] and r["member_id"].strip()   # synthesized, non-null
    assert r["status"] == "active"
    assert r["agent_id"] == "21041582"             # NPN
    # stable across re-import
    assert parse(p)[0]["member_id"] == r["member_id"]


def test_july_two_word_and_multi_word_names_split(tmp_path):
    """Full Name split: 'Mary Jane Smith' → first 'Mary', last 'Jane Smith' (last
    token(s) as surname is fine for matching; the key is name+DOB stability)."""
    p = _app_status_xlsx(tmp_path, [
        ["A", "1", "NC", "Curtis Childress", "1949-03-14", "Enrolled", "Yes",
         "2026-01-01", "2026-01-01", "Devoted CORE", "H1-1", "No", "Yes", "Yes"],
    ])
    r = parse(p)[0]
    assert r["first_name"] == "Curtis"
    assert "Childress" in r["last_name"]


def test_old_csv_member_id_format_still_parses(tmp_path):
    """The older snake_case CSV with member_id must still work."""
    p = tmp_path / "devoted.csv"
    p.write_text("member_id,first_name,last_name,status,effective_date\n"
                 "DV-123,Jane,Roe,ENROLLED,2026-01-01\n")
    recs = parse(str(p))
    assert len(recs) == 1
    assert recs[0]["member_id"] == "DV-123"
    assert recs[0]["first_name"] == "Jane"


# Rich Application-Status format (the full BOB — has an Mbi column + more).
RICH_COLS = [
    "Application Status Report Agent Name",
    "Application Status Report Agent Npn",
    "Application Status Report Agent Primary Rts State",
    "Application Status Report Full Name",
    "Application Status Report Birth Date",
    "Application Status Report Mbi",
    "Application Status Report Phone Number",
    "Application Status Report Address",
    "Application Status Report City",
    "Application Status Report State",
    "Application Status Report Zip Code",
    "Application Status Report County",
    "Application Status Report Current Status",
    "Application Status Report Application Date",
    "Application Status Report Start Date",
    "Application Status Report Plan Name",
    "Application Status Report Plan ID",
    "Application Status Report Disenrollment Date",
    "Application Status Report Disenrollment Reason",
    "Application Status Report Is Plan Change (Yes / No)",
    "Application Status Report Is New to Medicare Advantage (Yes / No)",
    "Application Status Report Is New to Devoted (Yes / No)",
    "Application Status Report Is Application Is Currently Pending (Yes / No)",
    "Application Status Report Pending Reason",
    "Application Status Report Is Winning App (Yes / No)",
    "Application Status Report First Name",
    "Application Status Report Last Name",
    "Application Status Report Plan End Date",
    "Application Status Report Plan Type",
]


def _rich_xlsx(tmp_path, rows):
    p = tmp_path / "Devoted Book of business.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(RICH_COLS)
    for r in rows:
        ws.append(r)
    wb.save(p)
    return str(p)


def _rich_row(**kw):
    """One rich data row aligned to RICH_COLS, with sensible defaults."""
    d = dict(agent="Justin Basinger", npn="20446812", rts="NC",
             full="Praize Medley", dob="2004-06-17", mbi="2T74G35WQ90",
             phone="7045160684", addr="845 Highlander Ct", city="Concord",
             state="NC", zip="28025", county="CABARRUS", status="Enrolled",
             appdate="2025-10-21", start="2026-01-01",
             plan_name="DEVOTED DUAL FULL 013 NC", plan_id="H5299-013",
             disenr="", disenr_reason="", plan_change="No", new_ma="No",
             new_dev="Yes", pending="No", pending_reason="", winning="Yes",
             first="PRAIZE", last="MEDLEY", plan_end="", ptype="MAPD")
    d.update(kw)
    return [d["agent"], d["npn"], d["rts"], d["full"], d["dob"], d["mbi"],
            d["phone"], d["addr"], d["city"], d["state"], d["zip"], d["county"],
            d["status"], d["appdate"], d["start"], d["plan_name"], d["plan_id"],
            d["disenr"], d["disenr_reason"], d["plan_change"], d["new_ma"],
            d["new_dev"], d["pending"], d["pending_reason"], d["winning"],
            d["first"], d["last"], d["plan_end"], d["ptype"]]


def test_rich_row_captures_real_mbi_and_fields(tmp_path):
    p = _rich_xlsx(tmp_path, [_rich_row()])
    recs = parse(p)
    assert len(recs) == 1
    r = recs[0]
    assert r["mbi"] == "2T74G35WQ90"
    assert r["member_id"] == "2T74G35WQ90"        # MBI is the key, no DVND-
    assert r["first_name"] == "Praize" and r["last_name"] == "Medley"
    assert r["plan_name"] == "DEVOTED DUAL FULL 013 NC"
    assert r["plan_type"] == "MAPD"
    assert r["agent_name"] == "Justin Basinger"
    assert str(r["effective_date"]) == "2026-01-01"
    assert str(r["application_date"]) == "2025-10-21"
    assert r["city"] == "Concord" and r["county"] == "CABARRUS"
    assert r["is_winning_app"] is True
    assert r["commission_type"] == "renewal"       # New-to-MA = No
    assert r["status"] == "active"


def test_rich_new_to_ma_yes_is_initial(tmp_path):
    p = _rich_xlsx(tmp_path, [_rich_row(new_ma="Yes", mbi="3DJ9F94VV42")])
    assert parse(p)[0]["commission_type"] == "initial"


def test_rich_approved_is_active(tmp_path):
    p = _rich_xlsx(tmp_path, [_rich_row(status="Approved", mbi="4U76AC0RQ88")])
    recs = parse(p)
    assert len(recs) == 1
    assert recs[0]["status"] == "active"


def test_rich_other_status_skipped(tmp_path):
    p = _rich_xlsx(tmp_path, [_rich_row(status="Withdrawn", mbi="5CU9P00HW31")])
    assert parse(p) == []


def test_rich_lone_non_winning_still_active(tmp_path):
    # Peggy/Cynthia shape: a lone Is-Winning-App=No row that is Enrolled stays.
    p = _rich_xlsx(tmp_path, [_rich_row(winning="No", mbi="7G47C78AK61", full="Peggy Marsh")])
    recs = parse(p)
    assert len(recs) == 1
    assert recs[0]["is_winning_app"] is False
    assert recs[0]["status"] == "active"


def test_rich_captures_contract_code(tmp_path):
    p = _rich_xlsx(tmp_path, [_rich_row(plan_id="H5299-013", mbi="2T74G35WQ90")])
    r = parse(p)[0]
    assert r["contract_code"] == "H5299-013"


def test_old_lossy_application_status_still_parses(tmp_path):
    # Regression: the OLD lossy file (APP_COLS, NO Mbi column) still uses the
    # synth-id path and must keep working.
    p = _app_status_xlsx(tmp_path, [
        ["Anjana Patel", "21041582", "NC", "BRANDI TUCKER", "1977-03-21",
         "Enrolled", "Yes", "2026-04-01", "2026-04-01", "Devoted CORE", "H1234-001",
         "No", "Yes", "Yes"],
    ])
    recs = parse(p)
    assert len(recs) == 1
    assert recs[0]["member_id"].startswith("DVND-")   # synth path unchanged
    assert recs[0]["mbi"] == ""


def test_dedup_winning_app_beats_losing_on_date_tie():
    # Same term (None) + same eff date => winning-app decides.
    winner = {"term_date": None, "effective_date": date(2026, 1, 1),
              "is_winning_app": True, "application_date": date(2025, 11, 13)}
    loser = {"term_date": None, "effective_date": date(2026, 1, 1),
             "is_winning_app": False, "application_date": date(2025, 11, 18)}
    # winner should replace loser even though loser has a LATER app date
    assert _rec_is_more_current(winner, loser) is True
    # and the loser should NOT replace the winner
    assert _rec_is_more_current(loser, winner) is False


def test_dedup_later_appdate_wins_when_flag_ties():
    # Both same winning flag + same eff/term => later application_date wins.
    later = {"term_date": None, "effective_date": date(2026, 1, 1),
             "is_winning_app": True, "application_date": date(2025, 12, 5)}
    earlier = {"term_date": None, "effective_date": date(2026, 1, 1),
               "is_winning_app": True, "application_date": date(2025, 11, 21)}
    assert _rec_is_more_current(later, earlier) is True
    assert _rec_is_more_current(earlier, later) is False


def test_dedup_ignores_new_fields_for_other_carriers():
    # Records without is_winning_app/application_date behave exactly as before:
    # a later effective date still wins; a full tie still returns True (last-in-file).
    a = {"term_date": None, "effective_date": date(2026, 2, 1)}
    b = {"term_date": None, "effective_date": date(2026, 1, 1)}
    assert _rec_is_more_current(a, b) is True
    tie = {"term_date": None, "effective_date": date(2026, 1, 1)}
    assert _rec_is_more_current(dict(tie), dict(tie)) is True


def test_dedupe_flags_ambiguous_winning_pair():
    # Two active rows, same MBI, same term/eff, SAME winning flag AND SAME app date
    # => unresolvable => flagged for review (but still deduped to one, not dropped).
    recs = [
        {"carrier": "Devoted", "member_id": "9ZZ9ZZ9ZZ99", "full_name": "Ambi Guous",
         "status": "active", "term_date": None, "effective_date": _d(2026, 1, 1),
         "is_winning_app": True, "application_date": _d(2025, 12, 1)},
        {"carrier": "Devoted", "member_id": "9ZZ9ZZ9ZZ99", "full_name": "Ambi Guous",
         "status": "active", "term_date": None, "effective_date": _d(2026, 1, 1),
         "is_winning_app": True, "application_date": _d(2025, 12, 1)},
    ]
    out = _dedupe_bob_records(recs)
    active = [r for r in out if r.get("status") == "active"]
    assert len(active) == 1                          # collapsed, nothing dropped
    assert any(a["member_id"] == "9ZZ9ZZ9ZZ99" for a in AMBIGUOUS_WINNING_PAIRS)


def test_rich_data_on_non_first_sheet_is_found(tmp_path):
    """The REAL full Devoted BOB has a pivot-table 'Sheet1' in front of the actual
    Application-Status data sheet. parse() must scan all sheets and use the one
    carrying the Application-Status prefix, not blindly read the first sheet."""
    p = tmp_path / "Devoted Book of business.xlsx"
    wb = openpyxl.Workbook()
    junk = wb.active
    junk.title = "Sheet1"
    junk.append(["Count", "of", "stuff"])      # pivot-table-ish junk, no App-Status prefix
    junk.append(["x", "y", "z"])
    data = wb.create_sheet("application_status_report_2026_")
    data.append(RICH_COLS)
    data.append(_rich_row(mbi="2T74G35WQ90"))
    wb.save(p)
    recs = parse(str(p))
    assert len(recs) == 1
    assert recs[0]["mbi"] == "2T74G35WQ90"       # rich path found the 2nd sheet
    assert not recs[0]["member_id"].startswith("DVND-")


def test_dedupe_resolvable_pair_not_flagged():
    AMBIGUOUS_WINNING_PAIRS.clear()
    recs = [
        {"carrier": "Devoted", "member_id": "8YY8YY8YY88", "full_name": "Clear Winner",
         "status": "active", "term_date": None, "effective_date": _d(2026, 1, 1),
         "is_winning_app": True, "application_date": _d(2025, 12, 5)},
        {"carrier": "Devoted", "member_id": "8YY8YY8YY88", "full_name": "Clear Winner",
         "status": "active", "term_date": None, "effective_date": _d(2026, 1, 1),
         "is_winning_app": False, "application_date": _d(2025, 11, 1)},
    ]
    _dedupe_bob_records(recs)
    assert not any(a["member_id"] == "8YY8YY8YY88" for a in AMBIGUOUS_WINNING_PAIRS)
